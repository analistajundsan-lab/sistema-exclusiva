"""SST — Fase 3 (futuro avancado).

Endpoints aditivos, sem alterar o que ja existe em routes_sst.py:

- GET /sst/alertas          -> alertas automaticos por reincidencia (condutor/veiculo/bloqueio)
- GET /sst/score-preditivo  -> score heuristico de risco por condutor/veiculo/unidade
- GET /sst/comparativo      -> comparativo mensal por unidade + ranking corporativo
- GET /sst/export.xlsx      -> exportacao executiva do dashboard em XLSX
- GET /sst/export.pdf       -> exportacao executiva do dashboard em PDF

Todos respeitam o escopo de unidade do usuario (apply_user_unit_scope) e o mesmo
gate de papel do modulo SST. Nenhum dado e mockado: as agregacoes saem das tabelas
reais (Sinistro, LiberacaoCondutor, SafetyVehicle).
"""

import io
from collections import defaultdict
from datetime import datetime, date as date_type, timedelta
from zoneinfo import ZoneInfo
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import apply_user_unit_scope, require_role
from models import (
    LiberacaoCondutor,
    LiberacaoStatus,
    SafetyVehicle,
    Sinistro,
    User,
    UserRole,
    get_db,
)

router = APIRouter(prefix="/sst", tags=["sst-fase3"])
BRASILIA_TZ = ZoneInfo("America/Sao_Paulo")

SST_ROLES = (
    UserRole.TECNICO_SEGURANCA,
    UserRole.ENGENHEIRO_SEGURANCA,
    UserRole.ADMIN,
)


# ── Helpers ───────────────────────────────────────────────────────────────────


def _today() -> date_type:
    return datetime.now(BRASILIA_TZ).date()


def _scoped_sinistros(db: Session, user: User, unit: Optional[str]):
    q = apply_user_unit_scope(db.query(Sinistro), Sinistro.unit, user)
    if unit:
        q = q.filter(Sinistro.unit == unit)
    return q


def _grav_int(value) -> Optional[int]:
    try:
        n = int(str(value).strip()[0])
        return n if 1 <= n <= 5 else None
    except (ValueError, IndexError, TypeError):
        return None


def _nivel(score: int) -> str:
    if score >= 70:
        return "critico"
    if score >= 40:
        return "alto"
    if score >= 20:
        return "medio"
    return "baixo"


def _nivel_reincidencia(total: int) -> str:
    if total >= 4:
        return "critico"
    if total >= 3:
        return "alto"
    return "medio"


# ── Alertas automaticos por reincidencia ──────────────────────────────────────


@router.get("/alertas")
async def sst_alertas(
    unit: Optional[str] = None,
    dias: int = Query(90, ge=7, le=730),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*SST_ROLES)),
):
    """Condutores/veiculos com 2+ sinistros na janela e condutores com bloqueio
    recorrente na liberacao. Cada alerta tem nivel (medio/alto/critico)."""
    today = _today()
    inicio = today - timedelta(days=dias)
    sinistros = (
        _scoped_sinistros(db, current_user, unit)
        .filter(Sinistro.data_ocorrencia >= inicio)
        .all()
    )

    cond = defaultdict(list)
    veic = defaultdict(list)
    for s in sinistros:
        if s.condutor_nome:
            cond[s.condutor_nome].append(s.data_ocorrencia)
        if s.prefixo:
            veic[s.prefixo].append(s.data_ocorrencia)

    def _build(group: dict, key_name: str) -> list[dict]:
        out = []
        for nome, datas in group.items():
            if len(datas) < 2:
                continue
            ultima = max(datas)
            out.append(
                {
                    key_name: nome,
                    "total": len(datas),
                    "ultima_data": ultima.isoformat() if ultima else None,
                    "nivel": _nivel_reincidencia(len(datas)),
                }
            )
        out.sort(key=lambda x: x["total"], reverse=True)
        return out

    condutores = _build(cond, "condutor")
    veiculos = _build(veic, "prefixo")

    # Bloqueios recorrentes na liberacao de condutor (nao_liberado / restricao)
    lib_q = apply_user_unit_scope(
        db.query(LiberacaoCondutor), LiberacaoCondutor.unit, current_user
    )
    if unit:
        lib_q = lib_q.filter(LiberacaoCondutor.unit == unit)
    bloqueios = lib_q.filter(
        LiberacaoCondutor.resultado.in_(
            [LiberacaoStatus.NAO_LIBERADO, LiberacaoStatus.LIBERADO_COM_RESTRICAO]
        )
    ).all()
    blk = defaultdict(lambda: {"total": 0, "categorias": set()})
    for lb in bloqueios:
        rec = blk[lb.condutor_nome]
        rec["total"] += 1
        if lb.categoria_bloqueio:
            rec["categorias"].add(lb.categoria_bloqueio)
    bloqueios_recorrentes = [
        {
            "condutor": nome,
            "total": rec["total"],
            "categorias": sorted(rec["categorias"]),
            "nivel": _nivel_reincidencia(rec["total"]),
        }
        for nome, rec in blk.items()
        if rec["total"] >= 2
    ]
    bloqueios_recorrentes.sort(key=lambda x: x["total"], reverse=True)

    return {
        "window_days": dias,
        "condutores": condutores,
        "veiculos": veiculos,
        "bloqueios_recorrentes": bloqueios_recorrentes,
        "total_alertas": len(condutores) + len(veiculos) + len(bloqueios_recorrentes),
    }


# ── Score preditivo de risco ──────────────────────────────────────────────────


@router.get("/score-preditivo")
async def sst_score_preditivo(
    unit: Optional[str] = None,
    dias: int = Query(180, ge=30, le=730),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*SST_ROLES)),
):
    """Score heuristico 0-100 de propensao a risco por condutor, veiculo e unidade.

    Heuristico (NAO e modelo estatistico): combina frequencia de sinistros,
    gravidade media, presenca de vitima/afastamento e recencia (sinistros nos
    ultimos 30 dias pesam mais). Serve para priorizar atencao preventiva.
    """
    today = _today()
    inicio = today - timedelta(days=dias)
    sinistros = (
        _scoped_sinistros(db, current_user, unit)
        .filter(Sinistro.data_ocorrencia >= inicio)
        .all()
    )

    def _acc():
        return {
            "sinistros": 0,
            "grav_soma": 0,
            "grav_n": 0,
            "com_vitima": 0,
            "com_afastamento": 0,
            "recentes": 0,
        }

    cond = defaultdict(_acc)
    veic = defaultdict(_acc)
    uni = defaultdict(_acc)

    def _add(acc, s):
        acc["sinistros"] += 1
        g = _grav_int(s.gravidade)
        if g:
            acc["grav_soma"] += g
            acc["grav_n"] += 1
        if s.houve_vitima:
            acc["com_vitima"] += 1
        if s.houve_afastamento:
            acc["com_afastamento"] += 1
        if s.data_ocorrencia and (today - s.data_ocorrencia).days <= 30:
            acc["recentes"] += 1

    for s in sinistros:
        if s.condutor_nome:
            _add(cond[s.condutor_nome], s)
        if s.prefixo:
            _add(veic[s.prefixo], s)
        if s.unit:
            _add(uni[s.unit], s)

    def _score(acc) -> tuple[int, float]:
        grav_media = (acc["grav_soma"] / acc["grav_n"]) if acc["grav_n"] else 0.0
        raw = (
            acc["sinistros"] * 14
            + grav_media * 6
            + acc["com_afastamento"] * 12
            + acc["com_vitima"] * 9
            + acc["recentes"] * 8
        )
        return min(100, round(raw)), round(grav_media, 1)

    def _rank(group: dict, key_name: str) -> list[dict]:
        out = []
        for nome, acc in group.items():
            score, grav_media = _score(acc)
            out.append(
                {
                    key_name: nome,
                    "score": score,
                    "nivel": _nivel(score),
                    "sinistros": acc["sinistros"],
                    "gravidade_media": grav_media,
                    "com_vitima": acc["com_vitima"],
                    "com_afastamento": acc["com_afastamento"],
                }
            )
        out.sort(key=lambda x: x["score"], reverse=True)
        return out

    return {
        "window_days": dias,
        "condutores": _rank(cond, "condutor")[:15],
        "veiculos": _rank(veic, "prefixo")[:15],
        "unidades": _rank(uni, "unidade"),
    }


# ── Comparativo mensal por unidade + ranking corporativo ──────────────────────


@router.get("/comparativo")
async def sst_comparativo(
    meses: int = Query(6, ge=2, le=24),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*SST_ROLES)),
):
    """Sinistros por unidade ao longo dos ultimos N meses + ranking corporativo
    (total e taxa por veiculo ativo). Escopado pela permissao do usuario."""
    today = _today()
    labels: list[str] = []
    y, m = today.year, today.month
    for i in range(meses - 1, -1, -1):
        mm, yy = m - i, y
        while mm <= 0:
            mm += 12
            yy -= 1
        labels.append(f"{yy:04d}-{mm:02d}")
    primeiro = labels[0]
    inicio = date_type(int(primeiro[:4]), int(primeiro[5:]), 1)

    sinistros = (
        _scoped_sinistros(db, current_user, None)
        .filter(Sinistro.data_ocorrencia >= inicio)
        .all()
    )

    por_unidade = defaultdict(lambda: {mo: 0 for mo in labels})
    total_unidade = defaultdict(int)
    for s in sinistros:
        if not s.data_ocorrencia or not s.unit:
            continue
        mo = f"{s.data_ocorrencia.year:04d}-{s.data_ocorrencia.month:02d}"
        if mo in labels:
            por_unidade[s.unit][mo] += 1
            total_unidade[s.unit] += 1

    # Frota ativa por unidade (para taxa por veiculo)
    veic_q = apply_user_unit_scope(
        db.query(SafetyVehicle).filter(SafetyVehicle.active.is_(True)),
        SafetyVehicle.unit,
        current_user,
    )
    frota = defaultdict(int)
    for v in veic_q.all():
        if v.unit:
            frota[v.unit] += 1

    unidades = []
    for u in sorted(set(list(por_unidade) + list(frota))):
        serie = por_unidade.get(u, {mo: 0 for mo in labels})
        total = total_unidade.get(u, 0)
        unidades.append(
            {
                "unidade": u,
                "por_mes": [{"mes": mo, "total": serie.get(mo, 0)} for mo in labels],
                "total": total,
                "media_mensal": round(total / meses, 1),
            }
        )

    ranking = []
    for u in unidades:
        nome = u["unidade"]
        f = frota.get(nome, 0)
        taxa = round(u["total"] / f, 2) if f else 0.0
        ranking.append(
            {
                "unidade": nome,
                "total": u["total"],
                "frota_ativa": f,
                "taxa_por_veiculo": taxa,
            }
        )
    ranking.sort(key=lambda x: x["total"], reverse=True)

    return {"meses": labels, "unidades": unidades, "ranking": ranking}


# ── Exportacao executiva (XLSX / PDF) ─────────────────────────────────────────


async def _dashboard_data(db, current_user, unit, date_start, date_end) -> dict:
    """Reaproveita exatamente o calculo do dashboard-v2 (sem duplicar regra)."""
    from routes_sst import sst_dashboard_v2

    return await sst_dashboard_v2(
        unit=unit,
        date_start=date_start,
        date_end=date_end,
        db=db,
        current_user=current_user,
    )


def _period_label(data: dict) -> str:
    p = data.get("period", {})
    ini = (p.get("start") or "").split("-")
    fim = (p.get("end") or "").split("-")
    f = lambda x: "/".join(reversed(x)) if len(x) == 3 else ""
    return f"{f(ini)} a {f(fim)}"


@router.get("/export.xlsx")
async def export_xlsx(
    unit: Optional[str] = None,
    date_start: Optional[date_type] = None,
    date_end: Optional[date_type] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*SST_ROLES)),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = await _dashboard_data(db, current_user, unit, date_start, date_end)
    s = data["summary"]

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1E3A8A")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13)

    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Cockpit SST — Resumo executivo"
    ws["A1"].font = title_font
    ws["A2"] = f"Periodo: {_period_label(data)}"
    ws["A3"] = f"Unidade: {unit or 'Todas'}"
    kpis = [
        ("Indice de atencao", s["risk_score"]),
        ("Sinistros (periodo)", s["sinistros_periodo"]),
        ("Variacao vs anterior (%)", s["sinistros_delta_pct"]),
        ("Conformidade check-list (%)", s["checklist_compliance_pct"]),
        ("Condutores bloqueados", s["condutores_bloqueados"]),
        ("Em investigacao", s["sinistros_investigacao"]),
        ("Ocorrencias SST", s["ocorrencias_sst"]),
        ("Custo total (R$)", s["custo_total"]),
        ("Acoes abertas", s["acoes_abertas"]),
        ("Acoes vencidas", s["acoes_vencidas"]),
        ("Com vitima", s["com_vitima"]),
        ("Com afastamento", s["com_afastamento"]),
    ]
    ws["A5"] = "Indicador"
    ws["B5"] = "Valor"
    for c in ("A5", "B5"):
        ws[c].fill = head_fill
        ws[c].font = head_font
    for i, (label, val) in enumerate(kpis, start=6):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16

    def _sheet(title, rows, headers):
        sh = wb.create_sheet(title)
        for j, h in enumerate(headers, start=1):
            cell = sh.cell(row=1, column=j, value=h)
            cell.fill = head_fill
            cell.font = head_font
            sh.column_dimensions[chr(64 + j)].width = 26
        for ri, row in enumerate(rows, start=2):
            for j, val in enumerate(row, start=1):
                sh.cell(row=ri, column=j, value=val)

    _sheet(
        "Por tipo",
        [(b["tipo"], b["total"]) for b in data["breakdowns"]["por_tipo"]],
        ["Tipo de sinistro", "Total"],
    )
    _sheet(
        "Rankings",
        [(c["nome"], c["total"]) for c in data["rankings"]["condutores"]],
        ["Condutor", "Sinistros"],
    )
    _sheet(
        "Plano de acao",
        [
            (
                a["numero"] or f"#{a['sinistro_id']}",
                a["unit"],
                a["responsavel"] or "",
                a["prazo"] or "",
                a["dias_atraso"],
                a["status_acao"],
            )
            for a in data["actions"]
        ],
        ["Sinistro", "Unidade", "Responsavel", "Prazo", "Dias atraso", "Status"],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cockpit-sst-{_today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export.pdf")
async def export_pdf(
    unit: Optional[str] = None,
    date_start: Optional[date_type] = None,
    date_end: Optional[date_type] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*SST_ROLES)),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    data = await _dashboard_data(db, current_user, unit, date_start, date_end)
    s = data["summary"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elems = [
        Paragraph("Cockpit SST — Resumo executivo", styles["Title"]),
        Paragraph(f"Periodo: {_period_label(data)}", styles["Normal"]),
        Paragraph(f"Unidade: {unit or 'Todas'}", styles["Normal"]),
        Spacer(1, 0.6 * cm),
    ]

    kpis = [
        ["Indicador", "Valor"],
        ["Indice de atencao", str(s["risk_score"])],
        ["Sinistros (periodo)", str(s["sinistros_periodo"])],
        ["Variacao vs anterior (%)", str(s["sinistros_delta_pct"])],
        ["Conformidade check-list (%)", str(s["checklist_compliance_pct"])],
        ["Condutores bloqueados", str(s["condutores_bloqueados"])],
        ["Em investigacao", str(s["sinistros_investigacao"])],
        ["Ocorrencias SST", str(s["ocorrencias_sst"])],
        ["Custo total (R$)", str(s["custo_total"])],
        ["Acoes vencidas", str(s["acoes_vencidas"])],
    ]
    t = Table(kpis, colWidths=[10 * cm, 6 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F3F4F6")],
                ),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elems.append(t)
    elems.append(Spacer(1, 0.6 * cm))

    if data["actions"]:
        elems.append(Paragraph("Plano de acao (mais atrasados)", styles["Heading2"]))
        rows = [["Sinistro", "Unidade", "Responsavel", "Atraso", "Status"]]
        for a in data["actions"][:12]:
            rows.append(
                [
                    a["numero"] or f"#{a['sinistro_id']}",
                    a["unit"] or "",
                    a["responsavel"] or "—",
                    f"{a['dias_atraso']}d" if a["dias_atraso"] else "—",
                    a["status_acao"],
                ]
            )
        ta = Table(rows, colWidths=[3.5 * cm, 3 * cm, 4.5 * cm, 2 * cm, 3 * cm])
        ta.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elems.append(ta)

    doc.build(elems)
    buf.seek(0)
    fname = f"cockpit-sst-{_today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
