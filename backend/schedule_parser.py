import re
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable

from openpyxl import load_workbook


@dataclass
class ParsedScheduleLine:
    unit: str
    prefix_code: str
    driver_name: str
    line_code: str
    direction: str
    client_name: str
    route_name: str
    start_time: str
    end_time: str
    source_sheet: str
    source_row: int
    source_col: int


UNIT_NAMES = {
    "jundiai": "Jundiai",
    "jundiaí": "Jundiai",
    "caieiras": "Caieiras",
    "santana": "Santana de Parnaiba",
}


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def only_numbers(value) -> str:
    return re.sub(r"[^0-9]", "", safe_str(value))


def pad_time(value: str) -> str:
    hour, minute = value.split(":")
    return f"{int(hour):02d}:{minute}"


def parse_time_range(value) -> tuple[str, str] | None:
    match = re.search(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", safe_str(value))
    if not match:
        return None
    return pad_time(match.group(1)), pad_time(match.group(2))


def detect_direction(value) -> str:
    text = safe_str(value).lower()
    if text.startswith("e/") or " entrada" in text:
        return "ENTRADA"
    if text.startswith("s/") or " saida" in text or " saída" in text:
        return "SAIDA"
    return ""


def clean_client(value) -> str:
    text = safe_str(value)
    text = re.sub(r"^\s*[eEsS]/\s*", "", text)
    text = re.sub(r"entrada|sa[ií]da", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s{2,}", " ", text).strip()
    return text


def unit_from_sheet(sheet_name: str) -> str:
    key = sheet_name.strip().lower()
    return UNIT_NAMES.get(key, sheet_name.strip())


def parse_schedule_workbook(content: bytes) -> list[ParsedScheduleLine]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parsed: list[ParsedScheduleLine] = []

    for worksheet in workbook.worksheets:
        if worksheet.title.upper() == "ATUALIZAR":
            parsed.extend(parse_linear_sheet(worksheet))
        elif worksheet.title.strip().lower() in UNIT_NAMES:
            parsed.extend(parse_block_sheet(worksheet))

    return parsed


def parse_linear_sheet(worksheet) -> Iterable[ParsedScheduleLine]:
    unit = ""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    out: list[ParsedScheduleLine] = []
    for index, row in enumerate(rows[1:], start=2):
        prefix_code = clean_numeric_text(row[0] if len(row) > 0 else "")
        driver_name = safe_str(row[1] if len(row) > 1 else "")
        line_code = clean_numeric_text(row[2] if len(row) > 2 else "")
        direction = normalize_direction(row[3] if len(row) > 3 else "")
        client_name = clean_client(row[4] if len(row) > 4 else "")
        route_name = safe_str(row[5] if len(row) > 5 else "")
        start_time = clean_time_text(row[6] if len(row) > 6 else "")
        end_time = clean_time_text(row[7] if len(row) > 7 else "")
        unit = unit_from_sheet(safe_str(row[8] if len(row) > 8 else unit))

        if not any(
            [prefix_code, driver_name, line_code, client_name, start_time, end_time]
        ):
            continue
        if not start_time or not end_time:
            continue

        out.append(
            ParsedScheduleLine(
                unit=unit,
                prefix_code=prefix_code,
                driver_name=driver_name,
                line_code=line_code,
                direction=direction,
                client_name=client_name,
                route_name=route_name,
                start_time=start_time,
                end_time=end_time,
                source_sheet=worksheet.title,
                source_row=index,
                source_col=1,
            )
        )
    return out


def parse_block_sheet(worksheet) -> Iterable[ParsedScheduleLine]:
    rows = list(worksheet.iter_rows(values_only=True))
    max_col = worksheet.max_column or 0
    unit = unit_from_sheet(worksheet.title)
    out: list[ParsedScheduleLine] = []

    row_index = 2
    while row_index <= len(rows) - 4:
        row = rows[row_index]
        prefix_code = clean_numeric_text(row[0] if len(row) > 0 else "")
        driver_name = safe_str(row[2] if len(row) > 2 else "")

        if driver_name:
            for col_index in range(4, max_col):
                value = row[col_index] if col_index < len(row) else ""
                parsed_time = parse_time_range(value)
                if not parsed_time:
                    continue

                client_original = (
                    rows[row_index + 1][col_index]
                    if col_index < len(rows[row_index + 1])
                    else ""
                )
                line_original = (
                    rows[row_index + 2][col_index]
                    if col_index < len(rows[row_index + 2])
                    else ""
                )
                route_original = (
                    rows[row_index + 3][col_index]
                    if col_index < len(rows[row_index + 3])
                    else ""
                )

                out.append(
                    ParsedScheduleLine(
                        unit=unit,
                        prefix_code=prefix_code,
                        driver_name=driver_name,
                        line_code=only_numbers(line_original),
                        direction=detect_direction(client_original),
                        client_name=clean_client(client_original),
                        route_name=safe_str(route_original),
                        start_time=parsed_time[0],
                        end_time=parsed_time[1],
                        source_sheet=worksheet.title,
                        source_row=row_index + 1,
                        source_col=col_index + 1,
                    )
                )
            row_index += 4
        else:
            row_index += 1

    return out


def clean_numeric_text(value) -> str:
    text = safe_str(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return only_numbers(text) or text


def clean_time_text(value) -> str:
    text = safe_str(value)
    match = re.match(r"^(\d{1,2}):(\d{2})", text)
    if not match:
        return ""
    return f"{int(match.group(1)):02d}:{match.group(2)}"


def normalize_direction(value) -> str:
    text = safe_str(value).upper()
    if text in ("SAÍDA", "SAIDA", "S"):
        return "SAIDA"
    if text in ("ENTRADA", "E"):
        return "ENTRADA"
    return text
