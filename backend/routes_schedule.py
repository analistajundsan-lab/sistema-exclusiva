import asyncio
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import List, Optional
from zoneinfo import ZoneInfo

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from auth import (
    apply_user_unit_scope,
    ensure_unit_access,
    get_current_user,
    require_role,
    user_allowed_units,
)
from cache import cache_get, cache_set, schedule_version
import events
from models import (
    AuditLog,
    ScheduleImport,
    ScheduleLine,
    ScheduleLineStatus,
    ScheduleNonOperation,
    User,
    UserRole,
    get_db,
)
from schedule_parser import parse_schedule_workbook
from schemas import (
    CountResponse,
    ScheduleImportResponse,
    ScheduleImportPreviewClient,
    ScheduleImportPreviewResponse,
    ScheduleImportPreviewUnit,
    ScheduleLineResponse,
    ScheduleLineStatusChange,
    ScheduleLineUpdate,
    ScheduleNonOperationCreate,
    ScheduleSummaryItem,
    ScheduleWhatsappResponse,
)
from turn_reference_data import EXCLUDED_TURN_LINES, TURN_REFERENCE

router = APIRouter(prefix="/schedule", tags=["schedule"])

MAX_IMPORT_BYTES = 8 * 1024 * 1024
BRASILIA_TZ = ZoneInfo("America/Sao_Paulo")
UNIT_SHEET_TITLES = {
    "Jundiai": "jundiai",
    "Caieiras": "caieiras",
    "Santana de Parnaiba": "santana",
}

# Clientes que, mesmo mapeados em TURN_REFERENCE, devem aparecer como CARD AVULSO
# proprio no dashboard (decisao operacional) — nao diluidos dentro dos turnos.
FORCE_AVULSO_CLIENTS = {"PLATLOG"}

# Linhas especificas que viram um CARD AVULSO proprio com nome customizado, mesmo
# mapeadas em TURN_REFERENCE. Ex.: SORTATION MARABRAZ (4997/4998) — tinham rotulo
# de turno fora do padrao e nao apareciam em card nenhum no painel.
FORCE_AVULSO_LINES = {
    "4997": "SORTATION MARABRAZ",
    "4998": "SORTATION MARABRAZ",
}


async def parse_upload_file(file: UploadFile):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=422, detail="Envie uma planilha .xlsx sem macros"
        )

    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413, detail="Arquivo muito grande. Limite atual: 8 MB"
        )

    try:
        parsed_lines = parse_schedule_workbook(content)
    except Exception as exc:
        raise HTTPException(
            status_code=422, detail="Nao foi possivel ler a planilha enviada"
        ) from exc

    if not parsed_lines:
        raise HTTPException(
            status_code=422, detail="Nenhuma linha de escala encontrada na planilha"
        )

    return content, parsed_lines


def normalized_filename(file: UploadFile) -> str:
    return (file.filename or "escala.xlsx").strip()


def latest_import_ids_for_date(db: Session, schedule_date: date) -> list[int]:
    latest_effective_date = (
        db.query(func.max(ScheduleImport.effective_date))
        .filter(ScheduleImport.effective_date <= schedule_date)
        .scalar()
    )
    if latest_effective_date is None:
        return []

    rows = (
        db.query(ScheduleImport.id)
        .filter(ScheduleImport.effective_date == latest_effective_date)
        .all()
    )
    return [row[0] for row in rows]


def apply_schedule_date_scope(db: Session, query, schedule_date: Optional[date] = None):
    if not schedule_date:
        return query

    active_import_ids = latest_import_ids_for_date(db, schedule_date)
    if active_import_ids:
        return query.filter(ScheduleLine.import_id.in_(active_import_ids))

    # Compatibilidade com escalas antigas gravadas antes do versionamento.
    return query.filter(ScheduleLine.schedule_date == schedule_date)


def non_operating_ids_for_date(db: Session, schedule_date: Optional[date]) -> set[int]:
    """IDs de linhas marcadas como 'nao opera' naquele dia (por dia, volta sozinho)."""
    if not schedule_date:
        return set()
    rows = (
        db.query(ScheduleNonOperation.schedule_line_id)
        .filter(ScheduleNonOperation.operation_date == schedule_date)
        .all()
    )
    return {row[0] for row in rows}


def apply_operation_visibility(
    db: Session,
    query,
    schedule_date: Optional[date] = None,
    include_inactive: bool = False,
    hide_non_operating: bool = False,
):
    """Esconde da operacao o que nao roda:
    - is_active=False: desativada por periodo (so o ADM ve, via include_inactive).
    - 'nao opera' naquele dia: some apenas quando hide_non_operating=True.
    """
    if not include_inactive:
        query = query.filter(ScheduleLine.is_active.isnot(False))
    if hide_non_operating and schedule_date:
        nonop_subq = db.query(ScheduleNonOperation.schedule_line_id).filter(
            ScheduleNonOperation.operation_date == schedule_date
        )
        query = query.filter(ScheduleLine.id.notin_(nonop_subq))
    return query


def build_import_warnings(parsed_lines) -> list[str]:
    warnings: list[str] = []
    missing_line = sum(1 for line in parsed_lines if not line.line_code)
    missing_direction = sum(1 for line in parsed_lines if not line.direction)
    missing_client = sum(1 for line in parsed_lines if not line.client_name)
    overnight = sum(1 for line in parsed_lines if line.end_time < line.start_time)

    if missing_line:
        warnings.append(f"{missing_line} registros sem numero de linha")
    if missing_direction:
        warnings.append(f"{missing_direction} registros sem sentido Entrada/Saida")
    if missing_client:
        warnings.append(f"{missing_client} registros sem cliente")
    if overnight:
        warnings.append(f"{overnight} registros viram o dia")

    return warnings


@router.post("/import/preview", response_model=ScheduleImportPreviewResponse)
async def preview_schedule_import(
    file: UploadFile = File(...),
    schedule_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    _, parsed_lines = await parse_upload_file(file)
    filename = normalized_filename(file)
    unit_counts = Counter(line.unit for line in parsed_lines)
    client_counts = Counter(line.client_name or "Sem cliente" for line in parsed_lines)

    # Coexistencia: outras escalas na MESMA vigencia com nome diferente vao
    # SOMAR (duplicar) em vez de substituir. Mesmo nome -> substitui no lugar.
    existing_other_files: list[str] = []
    will_replace = False
    if schedule_date:
        for existing in (
            db.query(ScheduleImport)
            .filter(ScheduleImport.effective_date == schedule_date)
            .all()
        ):
            if existing.filename == filename:
                will_replace = True
            elif existing.filename not in existing_other_files:
                existing_other_files.append(existing.filename)

    return ScheduleImportPreviewResponse(
        total=len(parsed_lines),
        units=[
            ScheduleImportPreviewUnit(unit=unit, total=total)
            for unit, total in sorted(unit_counts.items())
        ],
        clients=[
            ScheduleImportPreviewClient(client_name=client, total=total)
            for client, total in client_counts.most_common(8)
        ],
        warnings=build_import_warnings(parsed_lines),
        effective_date=schedule_date,
        existing_other_files=existing_other_files,
        will_replace=will_replace,
    )


def apply_filters(
    query,
    schedule_date: Optional[date] = None,
    unit: Optional[str] = None,
    client_name: Optional[str] = None,
    line_code: Optional[str] = None,
    driver_name: Optional[str] = None,
    prefix_code: Optional[str] = None,
    status: Optional[ScheduleLineStatus] = None,
):
    if schedule_date:
        query = query.filter(ScheduleLine.schedule_date == schedule_date)
    if unit:
        query = query.filter(ScheduleLine.unit.ilike(f"%{unit}%"))
    if client_name:
        query = query.filter(ScheduleLine.client_name.ilike(f"%{client_name}%"))
    if line_code:
        query = query.filter(ScheduleLine.line_code.ilike(f"%{line_code}%"))
    if driver_name:
        query = query.filter(ScheduleLine.driver_name.ilike(f"%{driver_name}%"))
    if prefix_code:
        query = query.filter(ScheduleLine.prefix_code.ilike(f"%{prefix_code}%"))
    if status:
        query = query.filter(ScheduleLine.status == status)
    return query


def apply_start_window(
    query,
    start_in_minutes: Optional[int],
    operation_date: Optional[date] = None,
    now_brt: Optional[datetime] = None,
):
    if start_in_minutes is None:
        return query

    now_brt = now_brt or datetime.now(BRASILIA_TZ)
    if now_brt.tzinfo is None:
        now_brt = now_brt.replace(tzinfo=BRASILIA_TZ)
    else:
        now_brt = now_brt.astimezone(BRASILIA_TZ)
    cutoff_brt = now_brt + timedelta(minutes=start_in_minutes)
    now_str = now_brt.strftime("%H:%M")
    cutoff_str = cutoff_brt.strftime("%H:%M")

    if operation_date and operation_date != now_brt.date():
        return query

    if now_brt.date() == cutoff_brt.date():
        return query.filter(
            ScheduleLine.start_time >= now_str,
            ScheduleLine.start_time <= cutoff_str,
        )

    from sqlalchemy import or_

    return query.filter(
        or_(
            ScheduleLine.start_time >= now_str,
            ScheduleLine.start_time <= cutoff_str,
        )
    )


def safe_xlsx_text(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def block_client_label(line: ScheduleLine) -> str:
    prefix = (
        "E/"
        if line.direction.upper().startswith("E")
        else "S/" if line.direction.upper().startswith("S") else ""
    )
    return f"{prefix} {line.client_name}".strip()


def write_block_schedule_sheet(ws, lines: list[ScheduleLine]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    vehicle_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=3, value="MOTORISTA")
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=3).fill = header_fill
    ws.cell(row=1, column=3).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 4

    next_row = 4
    fallback_rows: dict[tuple[str, str], int] = {}
    used_columns: set[int] = set()

    for line in sorted(
        lines,
        key=lambda item: (
            item.source_row or 99999,
            item.source_col or 99999,
            item.start_time,
            item.line_code,
        ),
    ):
        row = line.source_row or fallback_rows.setdefault(
            (line.prefix_code, line.driver_name),
            next_row,
        )
        if not line.source_row and row == next_row:
            next_row += 4
        col = line.source_col or 5
        used_columns.add(col)

        ws.cell(row=row, column=1, value=safe_xlsx_text(line.prefix_code))
        ws.cell(row=row, column=3, value=safe_xlsx_text(line.driver_name))
        ws.cell(row=row, column=col, value=f"{line.start_time} - {line.end_time}")
        ws.cell(row=row + 1, column=col, value=safe_xlsx_text(block_client_label(line)))
        ws.cell(row=row + 2, column=col, value=safe_xlsx_text(f"L - {line.line_code}"))
        ws.cell(row=row + 3, column=col, value=safe_xlsx_text(line.route_name or ""))

        for r in range(row, row + 4):
            for c in (1, 3, col):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=1).fill = vehicle_fill
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=col).font = Font(bold=True)

    for col in sorted(used_columns):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22


def normalize_line_code(value: str) -> str:
    return "".join(ch for ch in str(value or "").strip() if ch.isdigit())


def normalize_dashboard_client(value: str) -> str:
    text = (value or "").strip().upper()
    compact = text.replace(" ", "")
    if "LUXOTTICA" in text:
        return "LUXOTTICA"
    if "GARDNER" in text:
        return "GARDNER"
    if "SORTATION" in text:
        return "SORTATION"
    if "SP-02" in compact or "SP02" in compact:
        return "SP02"
    if "SP04" in compact or "SP14" in compact:
        return "SP04/14"
    if "SP06" in compact:
        return "SP06"
    if "PLAT" in text:
        return "PLATLOG"
    return text or "SEM CLIENTE"


def start_minutes(value: Optional[str]) -> Optional[int]:
    """'HH:MM' -> minutos do dia. None se vazio/invalido."""
    if not value or ":" not in value:
        return None
    try:
        h, m = value.split(":")[:2]
        return int(h) * 60 + int(m)
    except (ValueError, TypeError):
        return None


def turn_by_similar_time(
    start_min: int, direction: str, refs: list, k: int = 5
) -> Optional[str]:
    """Encaixa uma linha SEM turno mapeado (ex.: MERCADO LIVRE) no turno das
    linhas JA mapeadas com horario de inicio mais parecido, no MESMO sentido
    (entrada com entrada, saida com saida). Voto da maioria entre os k vizinhos
    mais proximos no horario; empate -> turno do vizinho mais proximo.
    `refs` = lista de (direction_upper, start_min, turno)."""
    direction = (direction or "").upper()
    cand = sorted((abs(m - start_min), turn) for d, m, turn in refs if d == direction)[
        :k
    ]
    if not cand:
        return None
    counts = Counter(turn for _, turn in cand)
    top = max(counts.values())
    winners = {t for t, c in counts.items() if c == top}
    if len(winners) == 1:
        return next(iter(winners))
    for _, turn in cand:  # empate: desempata pelo vizinho mais proximo
        if turn in winners:
            return turn
    return cand[0][1]


def empty_direction_stats() -> dict:
    return {
        "entrada": 0,
        "saida": 0,
        "confirmed_entrada": 0,
        "confirmed_saida": 0,
        "pending_entrada": 0,
        "pending_saida": 0,
        "total": 0,
        "_lines": set(),
    }


def add_direction_stats(
    bucket: dict, line: ScheduleLine, operation_date: Optional[date] = None
) -> None:
    direction = (line.direction or "").upper()
    is_confirmed = (
        status_for_operation_date(line, operation_date) == ScheduleLineStatus.CONFIRMADA
    )
    if direction == "ENTRADA":
        bucket["entrada"] += 1
        bucket["confirmed_entrada" if is_confirmed else "pending_entrada"] += 1
    elif direction == "SAIDA":
        bucket["saida"] += 1
        bucket["confirmed_saida" if is_confirmed else "pending_saida"] += 1
    bucket["total"] += 1
    bucket["_lines"].add(normalize_line_code(line.line_code))


def status_for_operation_date(
    line: ScheduleLine, operation_date: Optional[date] = None
) -> ScheduleLineStatus:
    if (
        operation_date
        and operation_date == datetime.now(BRASILIA_TZ).date()
        and line.status == ScheduleLineStatus.CONFIRMADA
        and line.confirmed_at
    ):
        confirmed_at = line.confirmed_at
        if confirmed_at.tzinfo is None:
            confirmed_at = confirmed_at.replace(tzinfo=timezone.utc)
        confirmed_date = confirmed_at.astimezone(BRASILIA_TZ).date()
        if confirmed_date != operation_date:
            return ScheduleLineStatus.PENDENTE
    return line.status


def line_response_for_operation_date(
    line: ScheduleLine,
    operation_date: Optional[date] = None,
    non_operating: bool = False,
) -> dict:
    return {
        "id": line.id,
        "schedule_date": line.schedule_date,
        "unit": line.unit,
        "prefix_code": line.prefix_code,
        "driver_name": line.driver_name,
        "line_code": line.line_code,
        "direction": line.direction,
        "client_name": line.client_name,
        "route_name": line.route_name,
        "start_time": line.start_time,
        "end_time": line.end_time,
        "status": status_for_operation_date(line, operation_date),
        "is_active": line.is_active,
        "non_operating": non_operating,
        "notes": line.notes,
        "confirmed_by": line.confirmed_by,
        "confirmed_at": line.confirmed_at,
        "source_sheet": line.source_sheet,
        "source_row": line.source_row,
        "source_col": line.source_col,
        "created_by": line.created_by,
        "created_at": line.created_at,
    }


def public_direction_stats(bucket: dict) -> dict:
    data = {key: value for key, value in bucket.items() if key != "_lines"}
    data["unique_lines"] = len(bucket["_lines"])
    return data


def build_unit_summary(rows, operation_date: Optional[date] = None) -> list[dict]:
    """Agrega contagens por unidade aplicando status_for_operation_date.

    O resumo precisa bater com a lista: contar o status cru do banco mostraria
    de manha, como "confirmadas", linhas confirmadas ONTEM (que a lista ja
    exibe como pendentes por causa do reset diario das 00:00 BRT)."""
    buckets: dict[str, dict] = {}
    for row in rows:
        bucket = buckets.setdefault(
            row.unit,
            {
                "unit": row.unit,
                "total": 0,
                "entrada": 0,
                "saida": 0,
                "pending": 0,
                "confirmed": 0,
                "changed": 0,
                "cancelled": 0,
            },
        )
        bucket["total"] += 1
        direction = (row.direction or "").upper()
        if direction == "ENTRADA":
            bucket["entrada"] += 1
        elif direction == "SAIDA":
            bucket["saida"] += 1
        effective = status_for_operation_date(row, operation_date)
        if effective == ScheduleLineStatus.PENDENTE:
            bucket["pending"] += 1
        elif effective == ScheduleLineStatus.CONFIRMADA:
            bucket["confirmed"] += 1
        elif effective == ScheduleLineStatus.ALTERADA:
            bucket["changed"] += 1
        elif effective == ScheduleLineStatus.CANCELADA:
            bucket["cancelled"] += 1
    return [buckets[unit] for unit in sorted(buckets)]


SUMMARY_ROW_COLUMNS = (
    ScheduleLine.unit,
    ScheduleLine.direction,
    ScheduleLine.status,
    ScheduleLine.confirmed_at,
)


@router.post(
    "/import",
    response_model=ScheduleImportResponse,
    status_code=status.HTTP_201_CREATED,
)
async def import_schedule(
    schedule_date: date,
    replace: bool = True,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    content, parsed_lines = await parse_upload_file(file)
    filename = normalized_filename(file)

    try:
        if replace:
            schedule_import = (
                db.query(ScheduleImport)
                .filter(
                    ScheduleImport.effective_date == schedule_date,
                    ScheduleImport.filename == filename,
                )
                .first()
            )
            if schedule_import:
                db.query(ScheduleLine).filter(
                    ScheduleLine.import_id == schedule_import.id
                ).delete()
            else:
                schedule_import = ScheduleImport(
                    effective_date=schedule_date,
                    filename=filename,
                    file_size=len(content),
                    rows_imported=0,
                    created_by=current_user.id,
                )
                db.add(schedule_import)
                db.flush()
        else:
            schedule_import = ScheduleImport(
                effective_date=schedule_date,
                filename=filename,
                file_size=len(content),
                rows_imported=0,
                created_by=current_user.id,
            )
            db.add(schedule_import)
            db.flush()

        schedule_import.file_size = len(content)
        schedule_import.rows_imported = len(parsed_lines)
        schedule_import.created_by = current_user.id

        for parsed in parsed_lines:
            db.add(
                ScheduleLine(
                    import_id=schedule_import.id,
                    schedule_date=schedule_date,
                    unit=parsed.unit,
                    prefix_code=parsed.prefix_code,
                    driver_name=parsed.driver_name,
                    line_code=parsed.line_code,
                    direction=parsed.direction,
                    client_name=parsed.client_name,
                    route_name=parsed.route_name,
                    start_time=parsed.start_time,
                    end_time=parsed.end_time,
                    source_sheet=parsed.source_sheet,
                    source_row=parsed.source_row,
                    source_col=parsed.source_col,
                    created_by=current_user.id,
                )
            )

        db.add(
            AuditLog(
                user_id=current_user.id,
                action="IMPORT",
                resource="schedule",
                resource_id=schedule_import.id,
                details=f"{len(parsed_lines)} linhas importadas para {schedule_date}; arquivo={filename}",
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return ScheduleImportResponse(
        imported=len(parsed_lines), replaced=replace, schedule_date=schedule_date
    )


@router.get("/lines/count", response_model=CountResponse)
def count_schedule_lines(
    schedule_date: Optional[date] = None,
    unit: Optional[str] = None,
    client_name: Optional[str] = None,
    line_code: Optional[str] = None,
    driver_name: Optional[str] = None,
    prefix_code: Optional[str] = None,
    status: Optional[ScheduleLineStatus] = None,
    start_time_gte: Optional[str] = None,
    start_time_lt: Optional[str] = None,
    start_in_minutes: Optional[int] = Query(None, ge=1, le=1440),
    include_inactive: bool = False,
    hide_non_operating: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = apply_filters(
        db.query(ScheduleLine),
        None,
        unit,
        client_name,
        line_code,
        driver_name,
        prefix_code,
        None if schedule_date and status else status,
    )
    query = apply_schedule_date_scope(db, query, schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive, hide_non_operating
    )
    if start_time_gte:
        query = query.filter(ScheduleLine.start_time >= start_time_gte)
    if start_time_lt:
        query = query.filter(ScheduleLine.start_time < start_time_lt)
    query = apply_start_window(query, start_in_minutes, schedule_date)
    if schedule_date and status:
        return {
            "total": sum(
                1
                for line in query.all()
                if status_for_operation_date(line, schedule_date) == status
            )
        }
    return {"total": query.count()}


@router.get("/lines", response_model=List[ScheduleLineResponse])
def list_schedule_lines(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    schedule_date: Optional[date] = None,
    unit: Optional[str] = None,
    client_name: Optional[str] = None,
    line_code: Optional[str] = None,
    driver_name: Optional[str] = None,
    prefix_code: Optional[str] = None,
    prefix_code_exact: Optional[str] = None,
    status: Optional[ScheduleLineStatus] = None,
    start_in_minutes: Optional[int] = Query(None, ge=1, le=1440),
    include_inactive: bool = False,
    hide_non_operating: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = apply_filters(
        db.query(ScheduleLine),
        None,
        unit,
        client_name,
        line_code,
        driver_name,
        prefix_code,
        None if schedule_date and status else status,
    )
    # Match EXATO de prefixo (carro): o ilike parcial acima serve para busca,
    # mas nos paineis "outras linhas deste carro" (confirmar/trocar em lote) um
    # carro de prefixo curto (ex.: "3" da GARDNER em Jundiai) casava com 2230,
    # 2730 etc. e agrupava linhas de carros ERRADOS.
    if prefix_code_exact:
        query = query.filter(ScheduleLine.prefix_code == prefix_code_exact)
    query = apply_schedule_date_scope(db, query, schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive, hide_non_operating
    )
    query = apply_start_window(query, start_in_minutes, schedule_date)
    rows = query.order_by(
        ScheduleLine.unit, ScheduleLine.start_time, ScheduleLine.line_code
    ).all()
    if schedule_date and status:
        rows = [
            line
            for line in rows
            if status_for_operation_date(line, schedule_date) == status
        ]
    rows = rows[skip : skip + limit]
    if schedule_date:
        nonop_ids = non_operating_ids_for_date(db, schedule_date)
        return [
            line_response_for_operation_date(line, schedule_date, line.id in nonop_ids)
            for line in rows
        ]
    return rows


@router.patch("/lines/{line_id}", response_model=ScheduleLineResponse)
def update_schedule_line(
    line_id: int,
    body: ScheduleLineUpdate,
    db: Session = Depends(get_db),
    # Edicao inline tambem no painel de confirmacao (plantonista/analista),
    # alem dos gestores. Escopo de garagem garantido por ensure_unit_access.
    current_user: User = Depends(
        require_role(
            UserRole.ADMIN,
            UserRole.GERENTE,
            UserRole.SUPERVISAO,
            UserRole.SUPERVISOR,
            UserRole.ANALISTA,
            UserRole.PLANTONISTA,
        )
    ),
):
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    old_status = line.status
    update_data = body.model_dump(exclude_unset=True)
    new_status = update_data.get("status")

    # Mudar status via PATCH precisa manter as invariantes dos endpoints
    # dedicados (/confirm e /undo-confirm); sem isso, uma linha "confirmada"
    # sem confirmed_at nunca resetaria a meia-noite, e qualquer editor
    # reabriria confirmacao burlando o RBAC do undo.
    if new_status is not None and new_status != old_status:
        if old_status == ScheduleLineStatus.CONFIRMADA and not (
            getattr(current_user, "has_full_access", False)
            or current_user.role in (UserRole.ADMIN, UserRole.SUPERVISOR)
        ):
            raise HTTPException(
                status_code=403,
                detail="Reabrir linha confirmada exige Admin/Supervisor (Desfazer confirmacao)",
            )

    changes = []
    for field, value in update_data.items():
        old = getattr(line, field)
        if old != value:
            changes.append(f"{field}: {old} -> {value}")
            setattr(line, field, value)

    if new_status is not None and new_status != old_status:
        if new_status == ScheduleLineStatus.CONFIRMADA:
            line.confirmed_by = current_user.id
            line.confirmed_at = datetime.now(timezone.utc).replace(tzinfo=None)
        else:
            line.confirmed_by = None
            line.confirmed_at = None

    if changes and "status" not in update_data:
        line.status = ScheduleLineStatus.ALTERADA

    if changes:
        db.add(
            AuditLog(
                user_id=current_user.id,
                action="UPDATE",
                resource="schedule_line",
                resource_id=line.id,
                details="; ".join(changes)[:500],
            )
        )
    db.commit()
    db.refresh(line)
    return line


@router.post("/lines/{line_id}/confirm", response_model=ScheduleLineResponse)
def confirm_schedule_line(
    line_id: int,
    db: Session = Depends(get_db),
    # Confirmacao e da operacao (Trafego/Analista/Admin + legados); cargos SST
    # (TST/Engenheiro) nao confirmam escala. Escopo de garagem via ensure_unit_access.
    current_user: User = Depends(
        require_role(
            UserRole.ADMIN,
            UserRole.GERENTE,
            UserRole.SUPERVISAO,
            UserRole.SUPERVISOR,
            UserRole.ANALISTA,
            UserRole.PLANTONISTA,
            UserRole.OPERATOR,
        )
    ),
):
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)
    if line.status == ScheduleLineStatus.CANCELADA:
        raise HTTPException(
            status_code=422, detail="Linha cancelada nao pode ser confirmada"
        )
    if line.is_active is False:
        raise HTTPException(
            status_code=422, detail="Linha desativada nao pode ser confirmada"
        )
    today_brt = datetime.now(BRASILIA_TZ).date()
    if line.id in non_operating_ids_for_date(db, today_brt):
        raise HTTPException(
            status_code=422,
            detail="Linha marcada como 'nao opera hoje' nao pode ser confirmada",
        )
    # Ja confirmada HOJE: no-op idempotente. Evita sobrescrever o autor da
    # confirmacao e poluir a auditoria com CONFIRMs duplicados (duplo toque/
    # segundo usuario que ainda nao viu a tela atualizar). Linha confirmada SEM
    # confirmed_at (dado legado) nao entra no no-op: reconfirma para ganhar o
    # carimbo e voltar a resetar a meia-noite.
    if (
        line.confirmed_at is not None
        and status_for_operation_date(line, today_brt) == ScheduleLineStatus.CONFIRMADA
    ):
        return line

    line.status = ScheduleLineStatus.CONFIRMADA
    line.confirmed_by = current_user.id
    # Grava UTC "naive" de forma deterministica: a coluna e TIMESTAMP sem fuso e,
    # se passassemos um datetime aware, o Postgres converteria pelo timezone da
    # SESSAO antes de descartar o offset. Forcando naive-UTC, o dashboard sempre
    # le confirmed_at como UTC e o "reset diario" das confirmacoes cai exatamente
    # 00:00 BRT (ver status_for_operation_date), em qualquer servidor/banco.
    line.confirmed_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="CONFIRM",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} confirmada; prefixo={line.prefix_code}; unidade={line.unit}",
        )
    )
    db.commit()
    db.refresh(line)
    return line


@router.post("/lines/{line_id}/undo-confirm", response_model=ScheduleLineResponse)
def undo_confirm_schedule_line(
    line_id: int,
    body: ScheduleLineStatusChange | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPERVISOR, UserRole.ADMIN)),
):
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)
    if line.status != ScheduleLineStatus.CONFIRMADA:
        raise HTTPException(
            status_code=422, detail="Apenas linha confirmada pode ser reaberta"
        )

    reason = body.reason if body else None
    line.status = ScheduleLineStatus.PENDENTE
    line.confirmed_by = None
    line.confirmed_at = None
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="UNDO_CONFIRM",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} reaberta; motivo={reason or 'nao informado'}",
        )
    )
    db.commit()
    db.refresh(line)
    return line


@router.post("/lines/{line_id}/cancel", response_model=ScheduleLineResponse)
def cancel_schedule_line(
    line_id: int,
    body: ScheduleLineStatusChange | None = None,
    db: Session = Depends(get_db),
    # Operadores da Confirmação de Escala (Tráfego/Analista) podem desativar uma
    # linha que não vai rodar, além de Admin/Supervisor. Escopo de garagem
    # garantido por ensure_unit_access abaixo.
    current_user: User = Depends(
        require_role(
            UserRole.ADMIN,
            UserRole.SUPERVISOR,
            UserRole.ANALISTA,
            UserRole.PLANTONISTA,
        )
    ),
):
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    reason = body.reason if body else None
    line.status = ScheduleLineStatus.CANCELADA
    if reason:
        line.notes = (
            reason if not line.notes else f"{line.notes} | Cancelamento: {reason}"
        )
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="CANCEL",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} cancelada; motivo={reason or 'nao informado'}",
        )
    )
    db.commit()
    db.refresh(line)
    return line


@router.post("/lines/{line_id}/deactivate", response_model=ScheduleLineResponse)
def deactivate_schedule_line(
    line_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Desativa a linha por periodo (ex.: Mercado Livre retirou a linha). Ela
    sai do painel de confirmacao/contagens ate ser reativada, sem ser apagada."""
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    line.is_active = False
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="DEACTIVATE",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} desativada (periodo); prefixo={line.prefix_code}; unidade={line.unit}",
        )
    )
    db.commit()
    db.refresh(line)
    return line


@router.post("/lines/{line_id}/reactivate", response_model=ScheduleLineResponse)
def reactivate_schedule_line(
    line_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Reativa uma linha desativada por periodo: volta ao painel de confirmacao."""
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    line.is_active = True
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="REACTIVATE",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} reativada; prefixo={line.prefix_code}; unidade={line.unit}",
        )
    )
    db.commit()
    db.refresh(line)
    return line


@router.get("/lines/{line_id}/pair", response_model=List[ScheduleLineResponse])
def schedule_line_pair(
    line_id: int,
    operation_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Linhas-par: mesma linha + mesma unidade (ex.: a Saida da Entrada
    selecionada). Usado pelo 'Nao operar' para marcar Entrada e Saida juntas."""
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)
    if not line.line_code:
        return []

    query = db.query(ScheduleLine).filter(
        ScheduleLine.id != line.id,
        ScheduleLine.unit == line.unit,
        ScheduleLine.line_code == line.line_code,
    )
    query = apply_schedule_date_scope(db, query, operation_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = query.filter(ScheduleLine.is_active.isnot(False))
    siblings = query.order_by(ScheduleLine.direction, ScheduleLine.start_time).all()

    nonop_ids = non_operating_ids_for_date(db, operation_date)
    return [
        line_response_for_operation_date(item, operation_date, item.id in nonop_ids)
        for item in siblings
    ]


@router.post("/lines/{line_id}/non-operation")
def set_schedule_non_operation(
    line_id: int,
    body: ScheduleNonOperationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_role(
            UserRole.ADMIN,
            UserRole.GERENTE,
            UserRole.SUPERVISAO,
            UserRole.SUPERVISOR,
            UserRole.ANALISTA,
            UserRole.PLANTONISTA,
        )
    ),
):
    """Marca que a(s) linha(s) NAO vao rodar apenas naquele dia. Some do painel
    e volta sozinha no dia seguinte. also_line_ids = linhas-par (ex.: a Saida)."""
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    target_ids = [line_id] + [i for i in body.also_line_ids if i != line_id]
    marked: list[int] = []
    for target_id in target_ids:
        target = db.query(ScheduleLine).filter(ScheduleLine.id == target_id).first()
        if not target:
            continue
        ensure_unit_access(current_user, target.unit)
        exists = (
            db.query(ScheduleNonOperation)
            .filter(
                ScheduleNonOperation.schedule_line_id == target_id,
                ScheduleNonOperation.operation_date == body.operation_date,
            )
            .first()
        )
        if not exists:
            # Savepoint: dois cliques simultaneos passam ambos no check acima;
            # o unique uq_nonop_line_date segura o segundo no banco e o
            # IntegrityError vira sucesso idempotente (nao um 500 na tela).
            try:
                with db.begin_nested():
                    db.add(
                        ScheduleNonOperation(
                            schedule_line_id=target_id,
                            operation_date=body.operation_date,
                            unit=target.unit,
                            line_code=target.line_code,
                            created_by=current_user.id,
                        )
                    )
                    db.flush()
            except IntegrityError:
                pass
        marked.append(target_id)

    db.add(
        AuditLog(
            user_id=current_user.id,
            action="NAO_OPERAR",
            resource="schedule_line",
            resource_id=line_id,
            details=f"Linhas {marked} nao operam em {body.operation_date}; unidade={line.unit}",
        )
    )
    db.commit()
    return {"marked": marked, "operation_date": body.operation_date}


@router.delete("/lines/{line_id}/non-operation", status_code=status.HTTP_204_NO_CONTENT)
def clear_schedule_non_operation(
    line_id: int,
    operation_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_role(
            UserRole.ADMIN,
            UserRole.GERENTE,
            UserRole.SUPERVISAO,
            UserRole.SUPERVISOR,
            UserRole.ANALISTA,
            UserRole.PLANTONISTA,
        )
    ),
):
    """Desfaz o 'nao operar' daquele dia: a linha volta a aparecer como pendente."""
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)

    # delete() em bulk NAO passa por session.deleted, entao o listener de
    # versao/SSE (models._track_schedule_changes) nao dispara e as telas dos
    # outros usuarios ficam presas no estado antigo. Deleta objeto a objeto.
    removals = (
        db.query(ScheduleNonOperation)
        .filter(
            ScheduleNonOperation.schedule_line_id == line_id,
            ScheduleNonOperation.operation_date == operation_date,
        )
        .all()
    )
    for removal in removals:
        db.delete(removal)
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="VOLTAR_OPERAR",
            resource="schedule_line",
            resource_id=line_id,
            details=f"Linha {line.line_code} volta a operar em {operation_date}; unidade={line.unit}",
        )
    )
    db.commit()


@router.get("/summary", response_model=List[ScheduleSummaryItem])
def schedule_summary(
    schedule_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(ScheduleLine)
    query = apply_schedule_date_scope(db, query, schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive=False, hide_non_operating=True
    )

    # Agrega em Python (nao em SQL) para aplicar o reset diario das 00:00 BRT
    # (status_for_operation_date) — o volume por dia e pequeno (centenas).
    rows = query.with_entities(*SUMMARY_ROW_COLUMNS).all()
    return [
        ScheduleSummaryItem(**item) for item in build_unit_summary(rows, schedule_date)
    ]


@router.get("/board")
def schedule_board(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    schedule_date: Optional[date] = None,
    unit: Optional[str] = None,
    client_name: Optional[str] = None,
    line_code: Optional[str] = None,
    driver_name: Optional[str] = None,
    prefix_code: Optional[str] = None,
    status: Optional[ScheduleLineStatus] = None,
    start_in_minutes: Optional[int] = Query(None, ge=1, le=1440),
    include_inactive: bool = False,
    hide_non_operating: bool = False,
    fresh: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Painel de escala em UMA requisicao: lines + total + summary.

    Junta o que antes eram 3 chamadas (/lines + /lines/count + /summary), cortando
    2/3 dos round-trips por refresh (grande ganho no celular). O resultado fica
    em cache de memoria por poucos segundos (ver cache.py). ?fresh=1 ignora o
    cache e o repopula — o front usa logo apos confirmar/trocar para QUEM agiu
    ver o resultado correto na hora (sem a linha confirmada reaparecer).
    """
    allowed = user_allowed_units(current_user)
    scope_key = "*" if allowed is None else ",".join(sorted(allowed))
    # A versao entra na chave: qualquer escrita de escala bumpa a versao e
    # invalida o board de TODOS os workers automaticamente (sem limpeza manual).
    cache_key = "schedule_board:" + "|".join(
        str(part)
        for part in (
            schedule_version(),
            scope_key,
            schedule_date or "",
            unit or "",
            client_name or "",
            line_code or "",
            driver_name or "",
            prefix_code or "",
            status.value if status else "",
            start_in_minutes or "",
            int(include_inactive),
            int(hide_non_operating),
            skip,
            limit,
        )
    )
    if not fresh:
        cached = cache_get(cache_key)
        if cached is not None:
            return cached

    # ---- lines + total (espelha /lines e /lines/count) ----
    query = apply_filters(
        db.query(ScheduleLine),
        None,
        unit,
        client_name,
        line_code,
        driver_name,
        prefix_code,
        None if schedule_date and status else status,
    )
    query = apply_schedule_date_scope(db, query, schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive, hide_non_operating
    )
    query = apply_start_window(query, start_in_minutes, schedule_date)
    rows = query.order_by(
        ScheduleLine.unit, ScheduleLine.start_time, ScheduleLine.line_code
    ).all()
    if schedule_date and status:
        rows = [
            line
            for line in rows
            if status_for_operation_date(line, schedule_date) == status
        ]
    total = len(rows)
    page_rows = rows[skip : skip + limit]
    nonop_ids = non_operating_ids_for_date(db, schedule_date)
    lines_payload = [
        line_response_for_operation_date(line, schedule_date, line.id in nonop_ids)
        for line in page_rows
    ]

    # ---- summary (espelha /summary) ----
    summary_query = apply_schedule_date_scope(db, db.query(ScheduleLine), schedule_date)
    summary_query = apply_user_unit_scope(
        summary_query, ScheduleLine.unit, current_user
    )
    summary_query = apply_operation_visibility(
        db,
        summary_query,
        schedule_date,
        include_inactive=False,
        hide_non_operating=True,
    )
    # Mesma mascara de reset diario da lista acima: o card de resumo nao pode
    # contradizer os cards de linha (ex.: 06:00 com tudo "confirmado" de ontem).
    summary_rows = summary_query.with_entities(*SUMMARY_ROW_COLUMNS).all()
    summary_payload = build_unit_summary(summary_rows, schedule_date)

    payload = jsonable_encoder(
        {"lines": lines_payload, "total": total, "summary": summary_payload}
    )
    # TTL serve so como coletor de chaves de versoes antigas: o frescor de fato
    # vem da versao na chave (bump em toda escrita invalida o board na hora).
    cache_set(cache_key, payload, ttl_seconds=30)
    return payload


@router.get("/version")
async def schedule_version_endpoint(current_user: User = Depends(get_current_user)):
    """Versao atual da escala (inteiro que sobe a cada escrita). O front faz
    polling leve disto a cada ~2s e so recarrega o painel inteiro quando muda —
    tempo-real barato (~2s) sem baixar a escala a cada ciclo."""
    return {"v": schedule_version()}


@router.get("/events")
async def schedule_events(
    request: Request,
    current_user: User = Depends(get_current_user),
):
    """Stream SSE de mudancas de escala (tempo-real <1s). Empurra um evento
    sempre que alguem confirma/troca/cancela/etc — o front recarrega o painel na
    hora. E ADITIVO: se a conexao cair, o polling de versao (~2s) cobre.

    Eventos: linhas `data: {"unit": ..., "schedule_date": ...}`. Comentarios
    `: keep-alive` a cada 15s mantem a conexao viva atraves de proxies.
    """
    queue = events.subscribe()

    async def gen():
        try:
            yield ": ok\n\n"
            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = await asyncio.wait_for(queue.get(), timeout=15)
                    yield f"data: {msg}\n\n"
                except asyncio.TimeoutError:
                    yield ": keep-alive\n\n"
        finally:
            events.unsubscribe(queue)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@router.get("/dashboard-turns")
def schedule_dashboard_turns(
    schedule_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = apply_schedule_date_scope(db, db.query(ScheduleLine), schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive=False, hide_non_operating=True
    )
    lines = query.order_by(
        ScheduleLine.unit, ScheduleLine.start_time, ScheduleLine.line_code
    ).all()

    turn_order = ["T1", "T2", "T3", "T4", "T5", "APRENDIZ"]
    standard_turns = set(turn_order)
    units: dict[str, dict] = {}
    client_index: dict[str, dict] = defaultdict(empty_direction_stats)
    excluded_found: dict[str, dict] = defaultdict(empty_direction_stats)

    # Referencia horario->turno por unidade, a partir das linhas JA mapeadas em
    # TURN_REFERENCE (so turnos padrao). Usada para encaixar o MERCADO LIVRE nos
    # turnos pela similaridade de horario (mesmo sentido), em vez de card avulso.
    turn_ref: dict[str, list] = defaultdict(list)
    for line in lines:
        lc = normalize_line_code(line.line_code)
        if not lc or lc in EXCLUDED_TURN_LINES:
            continue
        ref = TURN_REFERENCE.get(lc)
        sm = start_minutes(line.start_time)
        if (
            ref
            and sm is not None
            and ref["turn"] in standard_turns
            and ref["client"] not in FORCE_AVULSO_CLIENTS
        ):
            turn_ref[line.unit].append(
                ((line.direction or "").upper(), sm, ref["turn"])
            )

    for line in lines:
        line_code = normalize_line_code(line.line_code)
        if not line_code:
            continue

        unit_data = units.setdefault(
            line.unit,
            {
                "unit": line.unit,
                "turns": {turn: empty_direction_stats() for turn in turn_order},
                "client_cards": defaultdict(empty_direction_stats),
                "total": empty_direction_stats(),
            },
        )

        if line_code in EXCLUDED_TURN_LINES:
            client = normalize_dashboard_client(line.client_name)
            add_direction_stats(excluded_found[client], line, schedule_date)
            continue

        reference = TURN_REFERENCE.get(line_code)
        forced_client = FORCE_AVULSO_LINES.get(line_code)
        if forced_client:
            reference = (
                None  # linha vira card avulso c/ nome custom (ex.: SORTATION MARABRAZ)
            )
        elif reference and reference["client"] in FORCE_AVULSO_CLIENTS:
            forced_client = reference[
                "client"
            ]  # ex.: PLATLOG -> card avulso, fora dos turnos
            reference = None
        if reference:
            turn = reference["turn"]
            client = reference["client"]
            if turn not in unit_data["turns"]:
                unit_data["turns"][turn] = empty_direction_stats()
            add_direction_stats(unit_data["turns"][turn], line, schedule_date)
            add_direction_stats(client_index[client], line, schedule_date)
        else:
            client = forced_client or normalize_dashboard_client(line.client_name)
            sm = start_minutes(line.start_time)
            # MERCADO LIVRE nao tem turno mapeado: encaixa no turno das linhas com
            # horario mais parecido (mesmo sentido), em vez de virar card avulso.
            # Continua aparecendo no indice por cliente (detalhe). Outros avulsos
            # (GARDNER, SP02 sem ref, etc.) seguem como card avulso, como antes.
            ml_turn = (
                turn_by_similar_time(sm, line.direction, turn_ref.get(line.unit, []))
                if "MERCADO LIVRE" in client and sm is not None
                else None
            )
            if ml_turn:
                if ml_turn not in unit_data["turns"]:
                    unit_data["turns"][ml_turn] = empty_direction_stats()
                add_direction_stats(unit_data["turns"][ml_turn], line, schedule_date)
            else:
                add_direction_stats(
                    unit_data["client_cards"][client], line, schedule_date
                )
            add_direction_stats(client_index[client], line, schedule_date)

        add_direction_stats(unit_data["total"], line, schedule_date)

    response_units = []
    for unit_name in sorted(units):
        unit_data = units[unit_name]
        response_units.append(
            {
                "unit": unit_name,
                "total": public_direction_stats(unit_data["total"]),
                "turns": [
                    {
                        "key": turn,
                        "label": "Aprendiz" if turn == "APRENDIZ" else turn,
                        **public_direction_stats(unit_data["turns"][turn]),
                    }
                    for turn in turn_order
                ],
                "client_cards": [
                    {
                        "client": client,
                        **public_direction_stats(stats),
                    }
                    for client, stats in sorted(unit_data["client_cards"].items())
                ],
            }
        )

    return {
        "schedule_date": schedule_date,
        "units": response_units,
        "client_index": [
            {
                "client": client,
                **public_direction_stats(stats),
            }
            for client, stats in sorted(client_index.items())
        ],
        "excluded": [
            {
                "client": client,
                **public_direction_stats(stats),
            }
            for client, stats in sorted(excluded_found.items())
        ],
    }


@router.get("/whatsapp", response_model=ScheduleWhatsappResponse)
def schedule_whatsapp_text(
    schedule_date: date,
    unit: str,
    only_changes: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = apply_schedule_date_scope(db, db.query(ScheduleLine), schedule_date).filter(
        ScheduleLine.unit == unit,
    )
    ensure_unit_access(current_user, unit)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive=False, hide_non_operating=True
    )
    if only_changes:
        query = query.filter(
            ScheduleLine.status.in_(
                [ScheduleLineStatus.ALTERADA, ScheduleLineStatus.CANCELADA]
            )
        )

    lines = query.order_by(ScheduleLine.start_time, ScheduleLine.line_code).all()
    # Cabecalho honesto com o conteudo: sem only_changes o texto lista a escala
    # INTEIRA — anunciar como "alteracoes" fazia o grupo entender que tudo mudou.
    if only_changes:
        header = [
            "ALTERACOES REALIZADAS NA ESCALA",
            f"Entram em vigor a partir do dia: {schedule_date.strftime('%d/%m/%Y')}",
            "",
            f"Unidade: {unit}",
            "",
        ]
    else:
        header = [
            "ESCALA DO DIA",
            f"Dia: {schedule_date.strftime('%d/%m/%Y')}",
            "",
            f"Unidade: {unit}",
            "",
        ]

    body = [
        f"- {line.start_time} as {line.end_time} | Linha {line.line_code} | {line.direction} | "
        f"{line.client_name} | Prefixo {line.prefix_code} | Motorista: {line.driver_name}"
        for line in lines
    ]
    if not body:
        body = ["- Nenhuma linha encontrada para os filtros informados."]

    return ScheduleWhatsappResponse(
        schedule_date=schedule_date,
        unit=unit,
        total=len(lines),
        text="\n".join(header + body),
    )


@router.delete("/lines/{line_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_schedule_line(
    line_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Linha de escala nao encontrada")
    ensure_unit_access(current_user, line.unit)
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="DELETE",
            resource="schedule_line",
            resource_id=line.id,
            details=f"Linha {line.line_code} excluida; prefixo={line.prefix_code}; unidade={line.unit}",
        )
    )
    db.delete(line)
    db.commit()


@router.get("/download")
def download_schedule(
    schedule_date: date,
    unit: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_role(UserRole.ADMIN, UserRole.GERENTE, UserRole.SUPERVISAO)
    ),
):
    """Gera planilha XLSX no formato em blocos, semelhante ao arquivo importado."""
    query = apply_schedule_date_scope(db, db.query(ScheduleLine), schedule_date)
    query = apply_user_unit_scope(query, ScheduleLine.unit, current_user)
    query = apply_operation_visibility(
        db, query, schedule_date, include_inactive=False, hide_non_operating=True
    )
    if unit:
        ensure_unit_access(current_user, unit)
        query = query.filter(ScheduleLine.unit == unit)
    lines = query.order_by(
        ScheduleLine.unit, ScheduleLine.start_time, ScheduleLine.line_code
    ).all()

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    selected_units = [unit] if unit else list(UNIT_SHEET_TITLES)
    for unit_name in selected_units:
        sheet_lines = [line for line in lines if line.unit == unit_name]
        if not sheet_lines and unit:
            continue
        ws = wb.create_sheet(UNIT_SHEET_TITLES.get(unit_name, unit_name.lower()))
        write_block_schedule_sheet(ws, sheet_lines)
    if not wb.worksheets:
        ws = wb.create_sheet(UNIT_SHEET_TITLES.get(unit or "Caieiras", "caieiras"))
        write_block_schedule_sheet(ws, [])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = schedule_date.strftime("%d-%m-%Y")
    filename = f"ESCALA GERAL {date_str} ATUALIZADA.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
