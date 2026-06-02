import json
from io import BytesIO
from datetime import date as date_type, datetime
from typing import List, Optional
from zoneinfo import ZoneInfo
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from auth import apply_user_unit_scope, ensure_unit_access, get_current_user, user_allowed_units
from models import AuditLog, ScheduleLine, User, VehicleChecklist, get_db
from schemas import ChecklistCreate, ChecklistResponse, ChecklistUpdate

router = APIRouter(prefix="/checklist", tags=["checklist"])

JSON_FIELDS = ["licenciamento", "checklist_colocado", "wifi_status", "evidencias"]
BRASILIA_TZ = ZoneInfo("America/Sao_Paulo")
REPORT_MEDIA_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.12"

STATUS_LABELS = {
    "FUNCIONAL": "Funcional",
    "VISITA_TECNICA": "Visita tecnica",
    "SIM_EM_DIA": "Sim - em dia",
    "VENCIDO": "Vencido",
    "NAO_LOCALIZADO": "Nao localizado",
    "SIM_LOCALIZADO": "Sim - localizado",
    "DANIFICADO": "Danificado - necessario troca",
    "SIM_REMOVIDO_COLOCADO_NOVO": "Sim - removido antigo e colocado novo",
    "EXTRAVIADO_COLOCADO_NOVO": "Extraviado - colocado novo",
    "NAO_MANUTENCAO_FORA_GARAGEM": "Nao - manutencao ou fora da garagem",
    "JA_POSSUI_CHECKLIST_MES": "Ja possui checklist do mes",
    "SEM_CHECKLIST_COLOCAR_NOVO": "Sem checklist - colocar novo",
    "SIM_FUNCIONAL": "Sim, funcional",
    "NAO_SEM_REDE": "Nao - conectado porem sem rede",
    "NAO_APARECE_LISTA": "Nao aparece na lista de Wi-Fi",
    "NAO_FUNCIONA_FRETADAO": "Nao funciona no app Fretadao",
    "SIM_VENCIDO": "Sim - vencido",
    "NAO_COLOCAR_NOVO": "Nao - colocar novo",
    "TEM": "Tem",
    "NAO_TEM": "Nao tem",
}


def _today_brasilia() -> date_type:
    return datetime.now(BRASILIA_TZ).date()


def _duplicate_today_query(db: Session, prefixo: str, garagem: Optional[str] = None):
    query = db.query(VehicleChecklist).filter(
        func.upper(VehicleChecklist.prefixo) == prefixo.strip().upper(),
        func.date(VehicleChecklist.created_at) == _today_brasilia(),
    )
    if garagem:
        query = query.filter(VehicleChecklist.garagem == garagem)
    return query


def _apply_situation_filter(query, situacao: Optional[str]):
    if not situacao:
        return query

    s = situacao.upper()
    wifi_problem = or_(
        VehicleChecklist.wifi_status.ilike("%NAO_SEM_REDE%"),
        VehicleChecklist.wifi_status.ilike("%NAO_APARECE_LISTA%"),
        VehicleChecklist.wifi_status.ilike("%NAO_FUNCIONA_FRETADAO%"),
        and_(VehicleChecklist.wifi_outro.isnot(None), VehicleChecklist.wifi_outro != ""),
    )
    doc_missing = or_(
        VehicleChecklist.crlv_status == "NAO_LOCALIZADO",
        VehicleChecklist.emtu_status == "NAO_LOCALIZADO",
        VehicleChecklist.artesp_status == "NAO_LOCALIZADO",
        VehicleChecklist.emdec_status == "NAO_LOCALIZADO",
        VehicleChecklist.bolsa_documentos == "NAO_TEM",
    )
    doc_expired = or_(
        VehicleChecklist.crlv_status == "VENCIDO",
        VehicleChecklist.artesp_status == "VENCIDO",
        VehicleChecklist.emdec_status == "VENCIDO",
    )
    camera_issue = or_(
        VehicleChecklist.camera_frontal == "VISITA_TECNICA",
        VehicleChecklist.camera_lateral_esq == "VISITA_TECNICA",
        VehicleChecklist.camera_lateral_dir == "VISITA_TECNICA",
        VehicleChecklist.camera_fadiga == "VISITA_TECNICA",
        VehicleChecklist.camera_ip_motorista == "VISITA_TECNICA",
        VehicleChecklist.camera_salao == "VISITA_TECNICA",
    )
    filters = {
        "WIFI_PROBLEMA": wifi_problem,
        "DOCUMENTO_FALTANDO": doc_missing,
        "DOCUMENTO_VENCIDO": doc_expired,
        "CAMERA_VISITA_TECNICA": camera_issue,
        "CRLV_FALTANDO": VehicleChecklist.crlv_status == "NAO_LOCALIZADO",
        "CRLV_VENCIDO": VehicleChecklist.crlv_status == "VENCIDO",
        "EMTU_FALTANDO": VehicleChecklist.emtu_status == "NAO_LOCALIZADO",
        "EMTU_DANIFICADO": VehicleChecklist.emtu_status == "DANIFICADO",
        "ARTESP_FALTANDO": VehicleChecklist.artesp_status == "NAO_LOCALIZADO",
        "ARTESP_VENCIDO": VehicleChecklist.artesp_status == "VENCIDO",
        "EMDEC_FALTANDO": VehicleChecklist.emdec_status == "NAO_LOCALIZADO",
        "EMDEC_VENCIDO": VehicleChecklist.emdec_status == "VENCIDO",
        "BOLSA_DOCUMENTOS_TEM": VehicleChecklist.bolsa_documentos == "TEM",
        "BOLSA_DOCUMENTOS_NAO_TEM": VehicleChecklist.bolsa_documentos == "NAO_TEM",
        "CHECKLIST_FISICO_PENDENTE": VehicleChecklist.checklist_colocado.ilike(
            "%SEM_CHECKLIST_COLOCAR_NOVO%"
        ),
    }
    return query.filter(filters[s]) if s in filters else query


def _apply_checklist_filters(
    query,
    current_user: User,
    prefixo: Optional[str] = None,
    garagem: Optional[str] = None,
    tipo: Optional[str] = None,
    situacao: Optional[str] = None,
    data_inicio: Optional[date_type] = None,
    data_fim: Optional[date_type] = None,
):
    query = apply_user_unit_scope(query, VehicleChecklist.garagem, current_user)

    if prefixo:
        query = query.filter(VehicleChecklist.prefixo.ilike(f"%{prefixo}%"))
    if garagem:
        ensure_unit_access(current_user, garagem)
        query = query.filter(VehicleChecklist.garagem == garagem)
    if tipo:
        query = query.filter(VehicleChecklist.tipo == tipo)
    query = _apply_situation_filter(query, situacao)
    if data_inicio:
        query = query.filter(func.date(VehicleChecklist.created_at) >= data_inicio)
    if data_fim:
        query = query.filter(func.date(VehicleChecklist.created_at) <= data_fim)
    return query


def _label(value):
    if value is None or value == "":
        return "NAO PREENCHIDO"
    if isinstance(value, bool):
        return "Sim" if value else "Nao"
    if isinstance(value, list):
        if not value:
            return "NAO PREENCHIDO"
        return " | ".join(STATUS_LABELS.get(str(item), str(item)) for item in value)
    return STATUS_LABELS.get(str(value), str(value))


def _report_status(c: VehicleChecklist) -> str:
    cam_values = [
        c.camera_frontal,
        c.camera_lateral_esq,
        c.camera_lateral_dir,
        c.camera_fadiga,
        c.camera_ip_motorista,
        c.camera_salao,
    ]
    if any(value == "VISITA_TECNICA" for value in cam_values):
        return "PENDENCIA"
    if c.crlv_status == "VENCIDO" or c.artesp_status == "VENCIDO" or c.emdec_status == "VENCIDO":
        return "PENDENCIA"
    if (
        c.crlv_status == "NAO_LOCALIZADO"
        or c.artesp_status == "NAO_LOCALIZADO"
        or c.emdec_status == "NAO_LOCALIZADO"
        or c.emtu_status in {"NAO_LOCALIZADO", "DANIFICADO"}
        or c.bolsa_documentos == "NAO_TEM"
    ):
        return "ATENCAO"
    wifi = _parse_json_list(c.wifi_status)
    if any(item != "SIM_FUNCIONAL" for item in wifi):
        return "ATENCAO"
    return "OK"


def _parse_json_list(value) -> list:
    if not value:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            data = json.loads(value)
            return data if isinstance(data, list) else []
        except Exception:
            return []
    return []


def _checklist_report_workbook(checklists: list[VehicleChecklist]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vistorias por carro"

    headers = [
        "ID",
        "Data/Hora",
        "Garagem",
        "Prefixo",
        "Tipo",
        "Auditor",
        "Situacao geral",
        "Camera frontal",
        "Camera lateral esquerda",
        "Camera lateral direita",
        "Camera fadiga",
        "Camera IP motorista",
        "Camera salao",
        "Leitor de embarque",
        "Ar condicionado",
        "CRLV",
        "EMTU QR code",
        "ARTESP",
        "EMDEC",
        "Checklist fisico",
        "Bolsa de documentos",
        "QR Code",
        "Adesivo leitor",
        "Placa senha Wi-Fi",
        "Wi-Fi status",
        "Wi-Fi descricao",
        "Observacoes",
        "Qtd. evidencias",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in checklists:
        created_at = c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else ""
        ws.append(
            [
                c.id,
                created_at,
                c.garagem,
                c.prefixo,
                c.tipo,
                c.auditor_name,
                _report_status(c),
                _label(c.camera_frontal),
                _label(c.camera_lateral_esq),
                _label(c.camera_lateral_dir),
                _label(c.camera_fadiga),
                _label(c.camera_ip_motorista),
                _label(c.camera_salao),
                _label(c.tem_leitor_embarque),
                _label(c.ar_condicionado),
                _label(c.crlv_status),
                _label(c.emtu_status),
                _label(c.artesp_status),
                _label(c.emdec_status),
                _label(_parse_json_list(c.checklist_colocado)),
                _label(c.bolsa_documentos),
                _label(c.qr_code),
                _label(c.adesivo_leitor),
                _label(c.placa_senha_wifi),
                _label(_parse_json_list(c.wifi_status)),
                _label(c.wifi_outro),
                _label(c.observacoes),
                len(_parse_json_list(c.evidencias)),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 8,
        "B": 18,
        "C": 16,
        "D": 12,
        "E": 14,
        "F": 24,
        "G": 16,
        "T": 38,
        "Y": 36,
        "Z": 32,
        "AA": 50,
    }
    for column in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=column).column_letter
        ws.column_dimensions[letter].width = widths.get(letter, 22)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    legend = wb.create_sheet("Legenda")
    legend.append(["Campo", "Como interpretar"])
    legend.append(["Cada linha", "Uma vistoria individual de um carro/prefixo."])
    legend.append(["NAO PREENCHIDO", "O item nao recebeu resposta no checklist."])
    legend.append(["Situacao geral OK", "Sem pendencias detectadas nos itens preenchidos."])
    legend.append(["Situacao geral ATENCAO", "Documento faltando/danificado, bolsa ausente ou problema de Wi-Fi."])
    legend.append(["Situacao geral PENDENCIA", "Camera em visita tecnica ou documento vencido."])
    for cell in legend[1]:
        cell.fill = header_fill
        cell.font = header_font
    legend.column_dimensions["A"].width = 26
    legend.column_dimensions["B"].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return _as_macro_enabled_package(output)


def _as_macro_enabled_package(xlsx_output: BytesIO) -> BytesIO:
    macro_output = BytesIO()
    xlsx_output.seek(0)
    with ZipFile(xlsx_output, "r") as source, ZipFile(
        macro_output, "w", ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8").replace(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
                data = text.encode("utf-8")
            target.writestr(item, data)
    macro_output.seek(0)
    return macro_output


def _to_response(c: VehicleChecklist) -> ChecklistResponse:
    data: dict = {col.name: getattr(c, col.name) for col in c.__table__.columns}
    for field in JSON_FIELDS:
        val = data.get(field)
        if isinstance(val, str):
            try:
                data[field] = json.loads(val)
            except Exception:
                data[field] = []
        elif val is None:
            data[field] = []
    return ChecklistResponse(**data)


@router.post("/", response_model=ChecklistResponse, status_code=status.HTTP_201_CREATED)
async def create_checklist(
    body: ChecklistCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ensure_unit_access(current_user, body.garagem)
    if _duplicate_today_query(db, body.prefixo, body.garagem).first():
        raise HTTPException(status_code=422, detail="CHECK-LIST REALIZADO HOJE")
    data = body.model_dump()
    for field in JSON_FIELDS:
        val = data.get(field)
        if isinstance(val, list):
            data[field] = json.dumps(val, ensure_ascii=False)
        else:
            data[field] = None

    checklist = VehicleChecklist(
        **data,
        auditor_id=current_user.id,
        auditor_name=current_user.display_name or current_user.name,
        created_at=datetime.now(BRASILIA_TZ).replace(tzinfo=None),
    )
    db.add(checklist)
    db.flush()
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="CREATE",
            resource="checklist",
            resource_id=checklist.id,
        )
    )
    db.commit()
    db.refresh(checklist)
    return _to_response(checklist)


@router.get("/garagens", response_model=List[str])
async def list_garagens(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna todas as garagens/unidades disponíveis no sistema."""
    rows_cl = db.query(VehicleChecklist.garagem).distinct().all()
    rows_sl = db.query(ScheduleLine.unit).distinct().all()
    garagens = sorted({r[0] for r in rows_cl + rows_sl if r[0]})
    allowed = user_allowed_units(current_user)
    if allowed is None:
        return garagens
    allowed_set = set(allowed)
    return [g for g in garagens if g in allowed_set]


@router.get("/exists-today")
async def checklist_exists_today(
    prefixo: str = Query(..., min_length=1),
    garagem: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if garagem:
        ensure_unit_access(current_user, garagem)
    query = _duplicate_today_query(db, prefixo, garagem)
    query = apply_user_unit_scope(query, VehicleChecklist.garagem, current_user)
    return {"exists": query.first() is not None}


@router.get("/", response_model=List[ChecklistResponse])
async def list_checklists(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    prefixo: Optional[str] = None,
    garagem: Optional[str] = None,
    tipo: Optional[str] = None,
    situacao: Optional[str] = None,
    data_inicio: Optional[date_type] = None,
    data_fim: Optional[date_type] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(VehicleChecklist).order_by(VehicleChecklist.created_at.desc())
    query = _apply_checklist_filters(
        query, current_user, prefixo, garagem, tipo, situacao, data_inicio, data_fim
    )

    return [_to_response(c) for c in query.offset(skip).limit(limit).all()]


@router.get("/download")
async def download_checklist_report(
    prefixo: Optional[str] = None,
    garagem: Optional[str] = None,
    tipo: Optional[str] = None,
    situacao: Optional[str] = None,
    data_inicio: Optional[date_type] = None,
    data_fim: Optional[date_type] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(VehicleChecklist).order_by(
        VehicleChecklist.garagem,
        VehicleChecklist.prefixo,
        VehicleChecklist.created_at.desc(),
    )
    query = _apply_checklist_filters(
        query, current_user, prefixo, garagem, tipo, situacao, data_inicio, data_fim
    )
    checklists = query.limit(5000).all()
    output = _checklist_report_workbook(checklists)

    filename_date = datetime.now(BRASILIA_TZ).strftime("%d-%m-%Y")
    filename = f"RELATORIO CHECKLIST {filename_date}.xlsm"
    return StreamingResponse(
        output,
        media_type=REPORT_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{checklist_id}", response_model=ChecklistResponse)
async def get_checklist(
    checklist_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    c = db.query(VehicleChecklist).filter(VehicleChecklist.id == checklist_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Checklist nao encontrado")
    ensure_unit_access(current_user, c.garagem)
    return _to_response(c)


@router.patch("/{checklist_id}", response_model=ChecklistResponse)
async def update_checklist(
    checklist_id: int,
    body: ChecklistUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin":
        raise HTTPException(
            status_code=403, detail="Apenas administradores podem editar checklists"
        )
    c = db.query(VehicleChecklist).filter(VehicleChecklist.id == checklist_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Checklist nao encontrado")
    ensure_unit_access(current_user, c.garagem)

    data = body.model_dump(exclude_unset=True)
    for field in JSON_FIELDS:
        if field in data:
            val = data[field]
            data[field] = (
                json.dumps(val, ensure_ascii=False) if isinstance(val, list) else None
            )

    for key, val in data.items():
        setattr(c, key, val)

    db.add(
        AuditLog(
            user_id=current_user.id,
            action="UPDATE",
            resource="checklist",
            resource_id=c.id,
        )
    )
    db.commit()
    db.refresh(c)
    return _to_response(c)


@router.delete("/{checklist_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_checklist(
    checklist_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role.value != "admin":
        raise HTTPException(
            status_code=403, detail="Apenas administradores podem excluir checklists"
        )
    c = db.query(VehicleChecklist).filter(VehicleChecklist.id == checklist_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Checklist nao encontrado")
    ensure_unit_access(current_user, c.garagem)
    db.add(
        AuditLog(
            user_id=current_user.id,
            action="DELETE",
            resource="checklist",
            resource_id=c.id,
        )
    )
    db.delete(c)
    db.commit()
