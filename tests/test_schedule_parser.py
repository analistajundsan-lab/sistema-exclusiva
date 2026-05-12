from io import BytesIO

from openpyxl import Workbook

from schedule_parser import parse_schedule_workbook


def build_block_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "caieiras"

    ws.cell(row=1, column=3, value="MOTORISTA")
    ws.cell(row=4, column=1, value="1580")
    ws.cell(row=4, column=3, value="E N DA SILVA")
    ws.cell(row=4, column=7, value="03:50 - 04:45")
    ws.cell(row=5, column=7, value="E/ M LIVRE - SP-02")
    ws.cell(row=6, column=7, value="L - 7368")
    ws.cell(row=7, column=7, value="JD. PINHEIROS / VERA TERESA")

    ws.cell(row=4, column=8, value="23:45 - 00:46")
    ws.cell(row=5, column=8, value="S/ MERCADO LIVRE")
    ws.cell(row=6, column=8, value="L - 7578")
    ws.cell(row=7, column=8, value="PQ. RESD. JUNDIAI")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_parse_block_schedule_to_linear_rows():
    rows = parse_schedule_workbook(build_block_workbook())

    assert len(rows) == 2

    first = rows[0]
    assert first.unit == "Caieiras"
    assert first.prefix_code == "1580"
    assert first.driver_name == "E N DA SILVA"
    assert first.line_code == "7368"
    assert first.direction == "ENTRADA"
    assert first.client_name == "M LIVRE - SP-02"
    assert first.route_name == "JD. PINHEIROS / VERA TERESA"
    assert first.start_time == "03:50"
    assert first.end_time == "04:45"

    second = rows[1]
    assert second.direction == "SAIDA"
    assert second.line_code == "7578"
    assert second.start_time == "23:45"
    assert second.end_time == "00:46"
