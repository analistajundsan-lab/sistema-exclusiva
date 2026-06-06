import pytest
from io import BytesIO
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import openpyxl
from main import app
from models import Base, User, Incident, Swap, get_db, UserRole
from auth import hash_password
import hashlib
import re

TEST_DB = "sqlite:///./test_phase2.db"
engine = create_engine(TEST_DB, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base.metadata.create_all(bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_teardown():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def hash_cpf(cpf: str) -> str:
    cpf_clean = re.sub(r"\D", "", cpf)
    return hashlib.sha256(cpf_clean.encode()).hexdigest()[:16]


@pytest.fixture
def auth_token():
    db = TestingSessionLocal()
    cpf_hash = hash_cpf("123.456.789-00")
    user = User(
        cpf_hash=cpf_hash,
        email="operator@test.com",
        name="Test Operator",
        password_hash=hash_password("password123"),
        role=UserRole.OPERATOR,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.close()

    response = client.post(
        "/auth/login", json={"cpf": "123.456.789-00", "password": "password123"}
    )
    return response.json()["access_token"]


@pytest.fixture
def supervisor_token():
    db = TestingSessionLocal()
    cpf_hash = hash_cpf("987.654.321-00")
    user = User(
        cpf_hash=cpf_hash,
        email="supervisor@test.com",
        name="Test Supervisor",
        password_hash=hash_password("password123"),
        role=UserRole.SUPERVISOR,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.close()

    response = client.post(
        "/auth/login", json={"cpf": "987.654.321-00", "password": "password123"}
    )
    return response.json()["access_token"]


@pytest.fixture
def admin_token():
    db = TestingSessionLocal()
    cpf_hash = hash_cpf("111.222.333-44")
    user = User(
        cpf_hash=cpf_hash,
        email="admin@test.com",
        name="Test Admin",
        password_hash=hash_password("password123"),
        role=UserRole.ADMIN,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.close()

    response = client.post(
        "/auth/login", json={"cpf": "111.222.333-44", "password": "password123"}
    )
    return response.json()["access_token"]


class TestIncidents:
    def test_create_incident(self, auth_token):
        response = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={
                "prefix_code": "VP-001",
                "incident_type": "Avaria",
                "description": "Motor não liga",
                "line": "Linha 501",
                "direction": "Centro",
            },
        )
        assert response.status_code == 201
        data = response.json()
        assert data["prefix_code"] == "VP-001"
        assert data["incident_type"] == "Avaria"

    def test_list_incidents(self, auth_token):
        client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        response = client.get(
            "/incidents/", headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0


class TestChecklist:
    def test_create_document_checklist_with_bolsa_documentos(self, auth_token):
        response = client.post(
            "/checklist/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={
                "garagem": "JUNDIAI",
                "prefixo": "1234",
                "tipo": "DOCUMENTOS",
                "crlv_status": "SIM_EM_DIA",
                "emtu_status": "SIM_LOCALIZADO",
                "artesp_status": "SIM_EM_DIA",
                "emdec_status": "SIM_EM_DIA",
                "bolsa_documentos": "TEM",
            },
        )

        assert response.status_code == 201
        data = response.json()
        assert data["tipo"] == "DOCUMENTOS"
        assert data["bolsa_documentos"] == "TEM"

    def test_block_duplicate_checklist_same_day(self, auth_token):
        payload = {
            "garagem": "JUNDIAI",
            "prefixo": "1234",
            "tipo": "AVULSO",
            "wifi_status": ["SIM_FUNCIONAL"],
        }

        first = client.post(
            "/checklist/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json=payload,
        )
        second = client.post(
            "/checklist/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json=payload,
        )

        assert first.status_code == 201
        assert second.status_code == 422
        assert second.json()["detail"] == "CHECK-LIST REALIZADO HOJE"

    def test_filter_checklists_by_wifi_problem(self, auth_token):
        headers = {"Authorization": f"Bearer {auth_token}"}
        client.post(
            "/checklist/",
            headers=headers,
            json={
                "garagem": "JUNDIAI",
                "prefixo": "1001",
                "tipo": "AVULSO",
                "wifi_status": ["SIM_FUNCIONAL"],
            },
        )
        client.post(
            "/checklist/",
            headers=headers,
            json={
                "garagem": "JUNDIAI",
                "prefixo": "1002",
                "tipo": "AVULSO",
                "wifi_status": ["NAO_SEM_REDE"],
            },
        )

        response = client.get(
            "/checklist/",
            headers=headers,
            params={"situacao": "WIFI_PROBLEMA"},
        )

        assert response.status_code == 200
        data = response.json()
        assert [item["prefixo"] for item in data] == ["1002"]

    def test_download_checklist_report_has_individual_vehicle_rows(self, auth_token):
        headers = {"Authorization": f"Bearer {auth_token}"}
        client.post(
            "/checklist/",
            headers=headers,
            json={
                "garagem": "JUNDIAI",
                "prefixo": "2001",
                "tipo": "DOCUMENTOS",
                "crlv_status": "NAO_LOCALIZADO",
                "bolsa_documentos": "NAO_TEM",
            },
        )

        response = client.get("/checklist/download", headers=headers)

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.ms-excel.sheet.macroEnabled.12"
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb["Vistorias por carro"]
        headers_row = [cell.value for cell in ws[1]]
        values = dict(zip(headers_row, [cell.value for cell in ws[2]]))
        assert values["Prefixo"] == "2001"
        assert values["Tipo"] == "DOCUMENTOS"
        assert values["CRLV"] == "Nao localizado"
        assert values["Bolsa de documentos"] == "Nao tem"

    def test_list_incidents_with_filter(self, auth_token):
        client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        response = client.get(
            "/incidents/?prefix_code=VP",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert response.status_code == 200

    def test_get_incident(self, auth_token):
        create_resp = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        incident_id = create_resp.json()["id"]

        response = client.get(
            f"/incidents/{incident_id}",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert response.status_code == 200
        assert response.json()["id"] == incident_id

    def test_update_incident(self, auth_token):
        create_resp = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        incident_id = create_resp.json()["id"]

        response = client.put(
            f"/incidents/{incident_id}",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"incident_type": "Acidente"},
        )
        assert response.status_code == 200
        assert response.json()["incident_type"] == "Acidente"

    def test_delete_incident_forbidden(self, auth_token):
        create_resp = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        incident_id = create_resp.json()["id"]

        response = client.delete(
            f"/incidents/{incident_id}",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert response.status_code == 403

    def test_delete_incident_supervisor(self, auth_token, supervisor_token):
        create_resp = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        incident_id = create_resp.json()["id"]

        response = client.delete(
            f"/incidents/{incident_id}",
            headers={"Authorization": f"Bearer {supervisor_token}"},
        )
        assert response.status_code == 204

    def test_create_incident_unauthorized(self):
        response = client.post(
            "/incidents/",
            json={
                "prefix_code": "VP-001",
                "incident_type": "Avaria",
                "description": "Motor não liga",
                "line": "Linha 501",
                "direction": "Centro",
            },
        )
        assert response.status_code in (401, 403)  # HTTPBearer retorna 403 sem token

    def test_list_incidents_filter_by_line(self, auth_token):
        client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={
                "prefix_code": "VP-001",
                "incident_type": "Avaria",
                "line": "Linha 501",
            },
        )
        response = client.get(
            "/incidents/?line=Linha 501",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["line"] == "Linha 501"

    def test_get_incident_not_found(self, auth_token):
        response = client.get(
            "/incidents/99999", headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 404

    def test_update_incident_forbidden_other_user(self, auth_token, supervisor_token):
        # Criar incident com operator
        create_resp = client.post(
            "/incidents/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"prefix_code": "VP-001", "incident_type": "Avaria"},
        )
        incident_id = create_resp.json()["id"]

        # Tentar atualizar com supervisor (não é dono nem admin)
        response = client.put(
            f"/incidents/{incident_id}",
            headers={"Authorization": f"Bearer {supervisor_token}"},
            json={"incident_type": "Acidente"},
        )
        assert response.status_code == 403


class TestSwaps:
    def test_create_swap(self, auth_token):
        response = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={
                "vehicle_out": "VP-001",
                "vehicle_in": "VP-002",
                "lines_covered": "Linha 501, 502",
            },
        )
        assert response.status_code == 201
        data = response.json()
        assert data["vehicle_out"] == "VP-001"
        assert data["vehicle_in"] == "VP-002"

    def test_create_swap_same_vehicles(self, auth_token):
        response = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-001"},
        )
        assert response.status_code == 422

    def test_list_swaps(self, auth_token):
        client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-002"},
        )
        response = client.get(
            "/swaps/", headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200

    def test_get_swap(self, auth_token):
        create_resp = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-002"},
        )
        swap_id = create_resp.json()["id"]

        response = client.get(
            f"/swaps/{swap_id}", headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        assert response.json()["id"] == swap_id

    def test_update_swap(self, auth_token):
        create_resp = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-002"},
        )
        swap_id = create_resp.json()["id"]

        response = client.put(
            f"/swaps/{swap_id}",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_in": "VP-003"},
        )
        assert response.status_code == 200

    def test_delete_swap_supervisor(self, auth_token, supervisor_token):
        create_resp = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-002"},
        )
        swap_id = create_resp.json()["id"]

        response = client.delete(
            f"/swaps/{swap_id}", headers={"Authorization": f"Bearer {supervisor_token}"}
        )
        assert response.status_code == 204

    def test_delete_swap_admin(self, auth_token, admin_token):
        create_resp = client.post(
            "/swaps/",
            headers={"Authorization": f"Bearer {auth_token}"},
            json={"vehicle_out": "VP-001", "vehicle_in": "VP-002"},
        )
        swap_id = create_resp.json()["id"]

        response = client.delete(
            f"/swaps/{swap_id}", headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 204


class TestAuth:
    def test_refresh_token_unauthorized(self):
        response = client.post("/auth/refresh", headers={})
        # Sem cookie nem header de refresh -> 401 (credencial ausente/invalida)
        assert response.status_code == 401
