from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from io import BytesIO
import hashlib
import re

import pytest
import openpyxl
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from auth import hash_password
from main import app
from models import Base, ScheduleImport, ScheduleLine, User, UserRole, get_db

TEST_DB = "sqlite:///./test_schedule.db"
engine = create_engine(TEST_DB, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base.metadata.create_all(bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_teardown():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def hash_cpf(cpf: str) -> str:
    cpf_clean = re.sub(r"\D", "", cpf)
    return hashlib.sha256(cpf_clean.encode()).hexdigest()[:16]


def create_token(cpf: str, role: UserRole) -> str:
    db = TestingSessionLocal()
    user = User(
        cpf_hash=hash_cpf(cpf),
        email=f"{role.value}@test.com",
        name=f"Test {role.value}",
        password_hash=hash_password("password123"),
        role=role,
        unit=None if role == UserRole.ADMIN else "Caieiras",
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.close()

    response = client.post("/auth/login", json={"cpf": cpf, "password": "password123"})
    return response.json()["access_token"]


@pytest.fixture
def admin_token():
    return create_token("111.222.333-44", UserRole.ADMIN)


@pytest.fixture
def operator_token():
    return create_token("123.456.789-00", UserRole.OPERATOR)


def db_count(model) -> int:
    db = TestingSessionLocal()
    try:
        return db.query(model).count()
    finally:
        db.close()


def build_schedule_file(
    line_code: str = "7368",
    start_time: str = "03:50",
    end_time: str = "04:45",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "caieiras"
    ws.cell(row=4, column=1, value="1580")
    ws.cell(row=4, column=3, value="E N DA SILVA")
    ws.cell(row=4, column=7, value=f"{start_time} - {end_time}")
    ws.cell(row=5, column=7, value="E/ M LIVRE - SP-02")
    ws.cell(row=6, column=7, value=f"L - {line_code}")
    ws.cell(row=7, column=7, value="JD. PINHEIROS / VERA TERESA")
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_admin_imports_schedule(admin_token):
    response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    assert response.json()["imported"] == 1

    list_response = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert list_response.status_code == 200
    data = list_response.json()
    assert len(data) == 1
    assert data[0]["unit"] == "Caieiras"
    assert data[0]["line_code"] == "7368"


def test_dashboard_turns_groups_by_unit_turn_and_direction(admin_token):
    import_response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file("7368"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 201

    response = client.get(
        "/schedule/dashboard-turns?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["schedule_date"] == "2026-04-13"
    assert data["units"][0]["unit"] == "Caieiras"
    assert data["units"][0]["total"]["entrada"] == 1
    assert data["units"][0]["total"]["saida"] == 0

    t1 = next(turn for turn in data["units"][0]["turns"] if turn["key"] == "T1")
    assert t1["entrada"] == 1
    assert t1["saida"] == 0
    assert t1["unique_lines"] == 1
    assert data["client_index"][0]["client"] == "SP02"


def test_admin_previews_schedule_without_saving(admin_token):
    response = client.post(
        "/schedule/import/preview",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 1
    assert data["units"] == [{"unit": "Caieiras", "total": 1}]
    assert data["clients"][0]["client_name"] == "M LIVRE - SP-02"

    lines = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()
    assert lines == []


def test_operator_cannot_preview_schedule(operator_token):
    response = client.post(
        "/schedule/import/preview",
        headers={"Authorization": f"Bearer {operator_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403


def test_operator_cannot_import_schedule(operator_token):
    response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {operator_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403


def test_import_rejects_macro_enabled_file(admin_token):
    response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsm",
                build_schedule_file(),
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )

    assert response.status_code == 422
    assert "xlsx" in response.json()["detail"]


def test_replace_only_affects_selected_date(admin_token):
    db = TestingSessionLocal()
    db.add(
        ScheduleLine(
            schedule_date=date(2026, 4, 12),
            unit="Caieiras",
            prefix_code="9999",
            driver_name="MOTORISTA ANTIGO",
            line_code="1111",
            direction="ENTRADA",
            client_name="M LIVRE - SP-02",
            start_time="03:00",
            end_time="04:00",
            created_by=1,
        )
    )
    db.commit()
    db.close()

    response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file("7467"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201

    old_day = client.get(
        "/schedule/lines?schedule_date=2026-04-12",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()
    new_day = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()

    assert len(old_day) == 1
    assert old_day[0]["line_code"] == "1111"
    assert len(new_day) == 1
    assert new_day[0]["line_code"] == "7467"


def test_same_file_and_effective_date_replaces_previous_import(admin_token):
    first = client.post(
        "/schedule/import?schedule_date=2026-05-01&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_01-05.xlsx",
                build_schedule_file("2828", "12:10", "13:10"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    second = client.post(
        "/schedule/import?schedule_date=2026-05-01&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_01-05.xlsx",
                build_schedule_file("2828", "12:00", "13:00"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert first.status_code == 201
    assert second.status_code == 201

    lines = client.get(
        "/schedule/lines?schedule_date=2026-05-01&line_code=2828",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()

    db = TestingSessionLocal()
    try:
        assert db.query(ScheduleImport).count() == 1
        assert db.query(ScheduleLine).count() == 1
    finally:
        db.close()

    assert len(lines) == 1
    assert lines[0]["start_time"] == "12:00"


def test_different_effective_dates_keep_history_and_query_active_version(admin_token):
    client.post(
        "/schedule/import?schedule_date=2026-05-01&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_vigencia_01-05.xlsx",
                build_schedule_file("2828", "12:10", "13:10"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    client.post(
        "/schedule/import?schedule_date=2026-05-20&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_vigencia_20-05.xlsx",
                build_schedule_file("2828", "12:00", "13:00"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    before_change = client.get(
        "/schedule/lines?schedule_date=2026-05-19&line_code=2828",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()
    after_change = client.get(
        "/schedule/lines?schedule_date=2026-05-21&line_code=2828",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()

    db = TestingSessionLocal()
    try:
        assert db.query(ScheduleImport).count() == 2
        assert db.query(ScheduleLine).count() == 2
    finally:
        db.close()

    assert len(before_change) == 1
    assert before_change[0]["schedule_date"] == "2026-05-01"
    assert before_change[0]["start_time"] == "12:10"
    assert len(after_change) == 1
    assert after_change[0]["schedule_date"] == "2026-05-20"
    assert after_change[0]["start_time"] == "12:00"


def test_different_files_same_effective_date_are_visible_together(admin_token):
    client.post(
        "/schedule/import?schedule_date=2026-05-20&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_caieiras_20-05.xlsx",
                build_schedule_file("2828", "12:00", "13:00"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    client.post(
        "/schedule/import?schedule_date=2026-05-20&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_extra_20-05.xlsx",
                build_schedule_file("3030", "14:00", "15:00"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    lines = client.get(
        "/schedule/lines?schedule_date=2026-05-21",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()

    assert db_count(ScheduleImport) == 2
    assert db_count(ScheduleLine) == 2
    assert {line["line_code"] for line in lines} == {"2828", "3030"}


def test_operator_receives_schedule_imported_by_admin(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-05-20&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala_geral_20-05.xlsx",
                build_schedule_file("2828", "12:00", "13:00"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    lines = client.get(
        "/schedule/lines?schedule_date=2026-05-21&unit=Caieiras",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    assert lines.status_code == 200
    assert len(lines.json()) == 1
    assert lines.json()[0]["line_code"] == "2828"


def test_start_window_uses_active_import_without_forcing_line_date(
    admin_token, operator_token
):
    now_brt = datetime.now(ZoneInfo("America/Sao_Paulo"))
    start_time = (now_brt + timedelta(minutes=10)).strftime("%H:%M")
    end_time = (now_brt + timedelta(minutes=50)).strftime("%H:%M")
    effective_date = now_brt.date() - timedelta(days=1)

    db = TestingSessionLocal()
    try:
        schedule_import = ScheduleImport(
            effective_date=effective_date,
            filename="escala_ativa.xlsx",
            rows_imported=1,
            created_by=1,
        )
        db.add(schedule_import)
        db.flush()
        db.add(
            ScheduleLine(
                import_id=schedule_import.id,
                schedule_date=effective_date,
                unit="Caieiras",
                prefix_code="1580",
                driver_name="MOTORISTA TESTE",
                line_code="2828",
                direction="ENTRADA",
                client_name="M LIVRE - SP-02",
                start_time=start_time,
                end_time=end_time,
                created_by=1,
            )
        )
        db.commit()
    finally:
        db.close()

    response = client.get(
        f"/schedule/lines?schedule_date={now_brt.date()}&status=pendente&start_in_minutes=40",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    count = client.get(
        f"/schedule/lines/count?schedule_date={now_brt.date()}&status=pendente&start_in_minutes=40",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    assert response.status_code == 200
    assert count.status_code == 200
    assert len(response.json()) == 1
    assert count.json()["total"] == 1
    assert response.json()[0]["line_code"] == "2828"


def test_confirm_schedule_line_moves_to_confirmed_history(admin_token, operator_token):
    import_response = client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 201

    pending = client.get(
        "/schedule/lines?schedule_date=2026-04-13&status=pendente",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    line_id = pending[0]["id"]

    confirm_response = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    assert confirm_response.status_code == 200
    assert confirm_response.json()["status"] == "confirmada"
    assert confirm_response.json()["confirmed_by"] is not None
    assert confirm_response.json()["confirmed_at"] is not None

    active_after = client.get(
        "/schedule/lines?schedule_date=2026-04-13&status=pendente",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    history_after = client.get(
        "/schedule/lines?schedule_date=2026-04-13&status=confirmada",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()

    assert active_after == []
    assert len(history_after) == 1
    assert history_after[0]["id"] == line_id


def test_schedule_board_consolidates_lines_total_summary(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    board = client.get(
        "/schedule/board?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    assert board.status_code == 200
    data = board.json()
    assert set(data.keys()) == {"lines", "total", "summary"}

    # Consistencia com os endpoints originais que o /board substitui.
    lines = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    count = client.get(
        "/schedule/lines/count?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    assert data["total"] == count["total"]
    assert len(data["lines"]) == len(lines)


def test_schedule_board_fresh_reflects_confirmation(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/board?schedule_date=2026-04-13&status=pendente",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["lines"][0]["id"]

    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    # fresh=1 ignora o cache: a linha confirmada nao volta para os pendentes.
    pending = client.get(
        "/schedule/board?schedule_date=2026-04-13&status=pendente&fresh=1",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    assert all(line["id"] != line_id for line in pending["lines"])


def test_schedule_version_bumps_on_confirm(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    v0 = client.get(
        "/schedule/version",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["v"]

    line_id = client.get(
        "/schedule/board?schedule_date=2026-04-13&status=pendente",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["lines"][0]["id"]
    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    v1 = client.get(
        "/schedule/version",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["v"]
    # Qualquer escrita de escala incrementa a versao global (tempo-real ~2s).
    assert v1 > v0


def test_schedule_events_requires_auth():
    # Sem token -> 401 (a rota existe e e protegida). NAO abrimos o stream
    # autenticado aqui porque ele e infinito e travaria o teste; o fluxo de
    # ponta-a-ponta do SSE e verificado por curl em producao.
    res = client.get("/schedule/events")
    assert res.status_code == 401


def test_confirm_schedule_line_is_idempotent(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]

    first = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    second = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json()["status"] == "confirmada"


def test_supervisor_can_cancel_line_and_operator_cannot(admin_token, operator_token):
    supervisor_token = create_token("222.333.444-55", UserRole.SUPERVISOR)
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]

    denied = client.post(
        f"/schedule/lines/{line_id}/cancel",
        headers={"Authorization": f"Bearer {operator_token}"},
        json={"reason": "Teste"},
    )
    assert denied.status_code == 403

    cancelled = client.post(
        f"/schedule/lines/{line_id}/cancel",
        headers={"Authorization": f"Bearer {supervisor_token}"},
        json={"reason": "Cliente cancelou"},
    )
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelada"
    assert "Cliente cancelou" in cancelled.json()["notes"]


def test_trafego_pode_desativar_linha(admin_token):
    """Trafego (plantonista) desativa uma linha que nao vai rodar -> cancelada."""
    trafego_token = create_token("555.666.777-88", UserRole.PLANTONISTA)
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {trafego_token}"},
    ).json()[0]["id"]

    cancelled = client.post(
        f"/schedule/lines/{line_id}/cancel",
        headers={"Authorization": f"Bearer {trafego_token}"},
        json={"reason": "Linha nao vai rodar"},
    )
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelada"


def test_admin_can_undo_confirmation(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    reopened = client.post(
        f"/schedule/lines/{line_id}/undo-confirm",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"reason": "Confirmado errado"},
    )
    assert reopened.status_code == 200
    data = reopened.json()
    assert data["status"] == "pendente"
    assert data["confirmed_by"] is None
    assert data["confirmed_at"] is None


def test_schedule_whatsapp_text_by_unit(admin_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    response = client.get(
        "/schedule/whatsapp?schedule_date=2026-04-13&unit=Caieiras",
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["unit"] == "Caieiras"
    assert data["total"] == 1
    # Sem only_changes o texto lista a escala inteira: o cabecalho nao pode
    # anunciar "alteracoes" (enganava o grupo do WhatsApp).
    assert "ESCALA DO DIA" in data["text"]
    assert "ALTERACOES REALIZADAS NA ESCALA" not in data["text"]
    assert "Linha 7368" in data["text"]
    assert "Prefixo 1580" in data["text"]

    only_changes = client.get(
        "/schedule/whatsapp?schedule_date=2026-04-13&unit=Caieiras&only_changes=true",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert only_changes.status_code == 200
    assert "ALTERACOES REALIZADAS NA ESCALA" in only_changes.json()["text"]


def test_download_xlsx_with_auth_header_after_schedule_change(admin_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()[0]

    update = client.patch(
        f"/schedule/lines/{line['id']}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"start_time": "04:10", "driver_name": "MOTORISTA ALTERADO"},
    )
    assert update.status_code == 200
    assert update.json()["status"] == "alterada"

    response = client.get(
        "/schedule/download?schedule_date=2026-04-13&unit=Caieiras",
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb["caieiras"]
    assert ws.cell(row=4, column=3).value == "MOTORISTA ALTERADO"
    assert ws.cell(row=4, column=7).value == "04:10 - 04:45"
    assert ws.cell(row=6, column=7).value == "L - 7368"


def test_audit_logs_are_listed(admin_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    response = client.get(
        "/audit/logs?resource=schedule&action=IMPORT",
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["resource"] == "schedule"
    assert data[0]["action"] == "IMPORT"


def test_swap_start_within_window():
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from routes_swaps import start_within_window

    tz = ZoneInfo("America/Sao_Paulo")
    now = datetime(2026, 6, 30, 12, 0, tzinfo=tz)
    # Dentro de +/- 180 min (turno do meio-dia).
    assert start_within_window("12:30", now, 180) is True
    assert start_within_window("10:00", now, 180) is True
    assert start_within_window("14:30", now, 180) is True
    # Fora da janela (turno da manha / da noite).
    assert start_within_window("03:40", now, 180) is False
    assert start_within_window("18:00", now, 180) is False
    # Virada de meia-noite: 00:15 esta a 45 min de 23:30.
    near_midnight = datetime(2026, 6, 30, 23, 30, tzinfo=tz)
    assert start_within_window("00:15", near_midnight, 60) is True
    # Sem horario (troca antiga) nao entra no envio por turno.
    assert start_within_window(None, now, 180) is False


def test_swap_can_be_created_from_confirmed_schedule_line(admin_token, operator_token):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    first_line = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]
    line_id = first_line["id"]
    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    response = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {operator_token}"},
        json={
            "schedule_line_id": line_id,
            "vehicle_out": "1580",
            "vehicle_in": "1590",
            "reason": "Adequacao operacional",
        },
    )

    assert response.status_code == 201
    data = response.json()
    assert data["schedule_line_id"] == line_id
    assert data["unit"] == "Caieiras"
    # Horario denormalizado da linha (usado no envio ao CCO por turno).
    assert data["start_time"] == first_line["start_time"]
    # Texto compacto: prefixo que atende (substituto) + linha 'E 7368'.
    assert "1590 - ATENDERA AS LINHAS" in data["whatsapp_text"]
    assert "E 7368" in data["whatsapp_text"]

    by_vehicle_out = client.get(
        "/swaps/?vehicle=1580",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    by_vehicle_in = client.get(
        "/swaps/?vehicle=1590",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    assert by_vehicle_out.status_code == 200
    assert by_vehicle_in.status_code == 200
    assert by_vehicle_out.json()[0]["id"] == data["id"]
    assert by_vehicle_in.json()[0]["id"] == data["id"]


def test_swap_can_change_only_driver_from_confirmed_schedule_line(
    admin_token, operator_token
):
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]
    client.post(
        f"/schedule/lines/{line['id']}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    response = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {operator_token}"},
        json={
            "schedule_line_id": line["id"],
            "vehicle_out": line["prefix_code"],
            "driver_in": "MOTORISTA RESERVA",
            "reason": "Adequacao operacional",
        },
    )

    assert response.status_code == 201
    data = response.json()
    assert data["vehicle_in"] is None
    assert data["driver_out"] == line["driver_name"]
    assert data["driver_in"] == "MOTORISTA RESERVA"
    # Troca so de motorista: mantem o prefixo e destaca o motorista.
    assert "1580" in data["whatsapp_text"]
    assert "MOT MOTORISTA RESERVA" in data["whatsapp_text"]


def test_swap_confirms_a_pending_schedule_line(admin_token, operator_token):
    # A troca de uma linha AINDA pendente e permitida e CONFIRMA a linha para o
    # dia (o carro previsto teve ocorrencia e outro vai rodar). So linha
    # CANCELADA nao pode ser trocada.
    client.post(
        "/schedule/import?schedule_date=2026-04-13&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]

    response = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {operator_token}"},
        json={"schedule_line_id": line_id, "vehicle_out": "1580", "vehicle_in": "1590"},
    )

    assert response.status_code == 201
    confirmed = client.get(
        "/schedule/lines?schedule_date=2026-04-13&status=confirmada",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    assert any(item["id"] == line_id for item in confirmed)


# --- Regressoes da auditoria de 2026-07-02 (reset diario, PATCH, dedup) ---


def today_brt_iso() -> str:
    return datetime.now(ZoneInfo("America/Sao_Paulo")).date().isoformat()


def import_schedule(admin_token, schedule_date: str = "2026-04-13") -> None:
    client.post(
        f"/schedule/import?schedule_date={schedule_date}&replace=true",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={
            "file": (
                "escala.xlsx",
                build_schedule_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def force_confirmed_yesterday(line_id: int) -> None:
    """Simula linha confirmada ONTEM (status cru CONFIRMADA no banco)."""
    from models import ScheduleLineStatus

    db = TestingSessionLocal()
    try:
        line = db.query(ScheduleLine).filter(ScheduleLine.id == line_id).first()
        line.status = ScheduleLineStatus.CONFIRMADA
        line.confirmed_at = datetime.now(timezone.utc).replace(
            tzinfo=None
        ) - timedelta(days=1)
        db.commit()
    finally:
        db.close()


def test_board_summary_applies_daily_reset(admin_token, operator_token):
    """Linha confirmada ontem conta como PENDENTE hoje tambem no summary.

    O resumo agregava o status cru do banco e, de manha, contradizia a lista
    (lista pendente x card '1 confirmada')."""
    import_schedule(admin_token)
    today = today_brt_iso()
    line_id = client.get(
        f"/schedule/board?schedule_date={today}",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["lines"][0]["id"]

    force_confirmed_yesterday(line_id)

    board = client.get(
        f"/schedule/board?schedule_date={today}&fresh=1",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    line = next(item for item in board["lines"] if item["id"] == line_id)
    assert line["status"] == "pendente"
    summary = next(item for item in board["summary"] if item["unit"] == "Caieiras")
    assert summary["pending"] == 1
    assert summary["confirmed"] == 0

    standalone = client.get(
        f"/schedule/summary?schedule_date={today}",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    item = next(entry for entry in standalone if entry["unit"] == "Caieiras")
    assert item["pending"] == 1
    assert item["confirmed"] == 0

    # Depois de confirmar HOJE, lista e resumo continuam de acordo.
    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    board = client.get(
        f"/schedule/board?schedule_date={today}&fresh=1",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()
    summary = next(item for item in board["summary"] if item["unit"] == "Caieiras")
    assert summary["confirmed"] == 1
    assert summary["pending"] == 0


def test_patch_status_confirmada_stamps_confirmation(admin_token):
    """Confirmar via PATCH mantem as invariantes do /confirm (carimbo de hoje);
    sem confirmed_at a linha nunca resetaria a meia-noite."""
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()[0]["id"]

    patched = client.patch(
        f"/schedule/lines/{line_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"status": "confirmada"},
    )
    assert patched.status_code == 200
    data = patched.json()
    assert data["status"] == "confirmada"
    assert data["confirmed_at"] is not None
    assert data["confirmed_by"] is not None


def test_patch_reopen_confirmed_line_requires_undo_role(admin_token, operator_token):
    """Voltar linha confirmada para pendente via PATCH nao pode burlar o RBAC
    do undo-confirm (Admin/Supervisor)."""
    plantonista_token = create_token("333.444.555-66", UserRole.PLANTONISTA)
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )

    denied = client.patch(
        f"/schedule/lines/{line_id}",
        headers={"Authorization": f"Bearer {plantonista_token}"},
        json={"status": "pendente"},
    )
    assert denied.status_code == 403

    allowed = client.patch(
        f"/schedule/lines/{line_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"status": "pendente"},
    )
    assert allowed.status_code == 200
    data = allowed.json()
    assert data["status"] == "pendente"
    assert data["confirmed_at"] is None
    assert data["confirmed_by"] is None


def test_confirm_keeps_first_author_when_reconfirmed_today(admin_token, operator_token):
    """Segundo usuario confirmando a mesma linha hoje e no-op: nao sobrescreve
    o autor nem duplica a trilha de auditoria."""
    plantonista_token = create_token("444.555.666-77", UserRole.PLANTONISTA)
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]

    first = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    second = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {plantonista_token}"},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json()["confirmed_by"] == first.json()["confirmed_by"]


def test_confirm_blocked_for_non_operating_today(admin_token, operator_token):
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    marked = client.post(
        f"/schedule/lines/{line_id}/non-operation",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"operation_date": today_brt_iso(), "also_line_ids": []},
    )
    assert marked.status_code == 200

    response = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    assert response.status_code == 422


def test_confirm_blocked_for_deactivated_line(admin_token, operator_token):
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    client.post(
        f"/schedule/lines/{line_id}/deactivate",
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    response = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {operator_token}"},
    )
    assert response.status_code == 422


def test_sst_roles_cannot_confirm_or_swap(admin_token, operator_token):
    tst_token = create_token("555.666.777-88", UserRole.TECNICO_SEGURANCA)
    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]

    confirm = client.post(
        f"/schedule/lines/{line_id}/confirm",
        headers={"Authorization": f"Bearer {tst_token}"},
    )
    assert confirm.status_code == 403

    swap = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {tst_token}"},
        json={"schedule_line_id": line_id, "vehicle_in": "1590"},
    )
    assert swap.status_code == 403


def test_version_bumps_when_non_operation_is_cleared(admin_token, operator_token):
    """Desfazer o 'nao opera' precisa bumpar a versao (e o SSE): o delete em
    bulk nao passava pelo listener e as telas dos outros ficavam presas."""
    import_schedule(admin_token)
    today = today_brt_iso()
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    client.post(
        f"/schedule/lines/{line_id}/non-operation",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"operation_date": today, "also_line_ids": []},
    )

    v0 = client.get(
        "/schedule/version",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["v"]
    cleared = client.delete(
        f"/schedule/lines/{line_id}/non-operation?operation_date={today}",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert cleared.status_code == 204
    v1 = client.get(
        "/schedule/version",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()["v"]
    assert v1 > v0


def test_swap_identical_resubmit_returns_existing(admin_token, operator_token):
    """Duplo-submit do PWA nao pode duplicar a troca (auditoria de producao
    achou trocas identicas criadas em dobro segundos depois)."""
    from models import Swap

    import_schedule(admin_token)
    line_id = client.get(
        "/schedule/lines?schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {operator_token}"},
    ).json()[0]["id"]
    payload = {"schedule_line_id": line_id, "vehicle_in": "1590"}

    first = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {operator_token}"},
        json=payload,
    )
    second = client.post(
        "/swaps/",
        headers={"Authorization": f"Bearer {operator_token}"},
        json=payload,
    )

    assert first.status_code == 201
    assert second.status_code == 201
    assert second.json()["id"] == first.json()["id"]
    assert db_count(Swap) == 1


def test_push_scan_uses_effective_schedule(admin_token, monkeypatch):
    """O scan do push precisa achar a escala pela VIGENCIA (nao por
    schedule_date == hoje, que nunca casava) e respeitar o dedup por dia."""
    import pytest as _pytest

    import push_service
    from config import settings
    from models import PushSentLine, PushSubscription

    brt = ZoneInfo("America/Sao_Paulo")
    now = datetime.now(brt)
    start = now + timedelta(minutes=5)
    if start.date() != now.date():
        _pytest.skip("janela de push cruza a meia-noite BRT")

    import_schedule(admin_token)
    db = TestingSessionLocal()
    try:
        line = db.query(ScheduleLine).first()
        line.start_time = start.strftime("%H:%M")
        db.add(
            PushSubscription(
                user_id=1,
                unit=line.unit,
                endpoint="https://push.example/sub-1",
                p256dh="key",
                auth="auth",
            )
        )
        db.commit()

        sent_payloads = []
        monkeypatch.setattr(settings, "VAPID_PUBLIC_KEY", "pub")
        monkeypatch.setattr(settings, "VAPID_PRIVATE_KEY", "priv")
        monkeypatch.setattr(settings, "PUSH_LEAD_MINUTES", 20)
        monkeypatch.setattr(
            push_service,
            "_send",
            lambda sub, payload: sent_payloads.append(payload) or True,
        )

        assert push_service.scan_and_notify(db) == 1
        assert len(sent_payloads) == 1
        assert db.query(PushSentLine).count() == 1
        # Segunda varredura do mesmo dia nao reenvia (dedup por linha/dia).
        assert push_service.scan_and_notify(db) == 0
    finally:
        db.close()


# --- Texto de trocas ao CCO: agrupado por prefixo, compacto, em ordem ---


def _seed_line(db, **kw):
    from models import ScheduleLineStatus

    defaults = dict(
        schedule_date=date(2026, 4, 13),
        unit="Caieiras",
        prefix_code="3380",
        driver_name="MOT",
        client_name="M LIVRE",
        end_time="23:59",
        status=ScheduleLineStatus.CONFIRMADA,
        is_active=True,
        created_by=1,
    )
    defaults.update(kw)
    line = ScheduleLine(**defaults)
    db.add(line)
    db.flush()
    return line


def _seed_swap(db, line, vehicle_in=None, driver_in=None):
    from models import Swap

    swap = Swap(
        schedule_line_id=line.id,
        schedule_date=line.schedule_date,
        unit=line.unit,
        vehicle_out=line.prefix_code,
        vehicle_in=vehicle_in,
        driver_in=driver_in,
        start_time=line.start_time,
        end_time=line.end_time,
        lines_covered=f"{line.direction} - {line.line_code}",
        created_by=1,
    )
    db.add(swap)
    db.flush()
    return swap


def test_swaps_text_groups_all_lines_of_prefix_in_schedule_order(admin_token):
    """Exemplo do Vinicius: prefixo 3380 atende 4 linhas. Mesmo registradas fora
    de ordem, saem juntas no mesmo texto, na sequencia da programacao."""
    db = TestingSessionLocal()
    try:
        l1 = _seed_line(db, line_code="4521", direction="SAIDA", start_time="13:45")
        l2 = _seed_line(db, line_code="5926", direction="SAIDA", start_time="15:50")
        l3 = _seed_line(db, line_code="7462", direction="ENTRADA", start_time="20:11")
        l4 = _seed_line(db, line_code="7412", direction="SAIDA", start_time="22:45")
        for line in (l3, l1, l4, l2):  # fora de ordem de proposito
            _seed_swap(db, line)
        db.commit()
    finally:
        db.close()

    res = client.get(
        "/swaps/whatsapp/text?unit=Caieiras&schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert res.status_code == 200
    text = res.json()["text"]
    assert "3380 - ATENDERA AS LINHAS : S 4521 - S 5926 - E 7462 - S 7412" in text


def test_swaps_text_uses_substitute_prefix_and_excludes_pending(admin_token):
    """Exemplo do Vinicius: 3400 substituido por 2300; so 7360 e 7431 trocadas.
    O texto sai sob 2300; a 7416 (sem troca) fica de fora."""
    db = TestingSessionLocal()
    try:
        a = _seed_line(
            db,
            prefix_code="3400",
            line_code="7360",
            direction="SAIDA",
            start_time="08:00",
        )
        b = _seed_line(
            db,
            prefix_code="3400",
            line_code="7431",
            direction="ENTRADA",
            start_time="09:00",
        )
        _seed_line(
            db,
            prefix_code="3400",
            line_code="7416",
            direction="SAIDA",
            start_time="10:00",
        )
        _seed_swap(db, a, vehicle_in="2300")
        _seed_swap(db, b, vehicle_in="2300")
        db.commit()
    finally:
        db.close()

    res = client.get(
        "/swaps/whatsapp/text?unit=Caieiras&schedule_date=2026-04-13",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert res.status_code == 200
    text = res.json()["text"]
    assert "2300 - ATENDERA AS LINHAS : S 7360 - E 7431" in text
    assert "7416" not in text
