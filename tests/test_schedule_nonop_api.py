"""Testes dos ajustes da escala:
- Desativar/Reativar por periodo (ADM) -> some/volta do painel via is_active.
- "Nao operar" por dia (plantonista) -> some so naquele dia, com par E/S.
- Edicao inline da linha pelo plantonista no painel.
"""

from io import BytesIO
import hashlib
import re

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from auth import hash_password
from main import app
from models import Base, User, UserRole, get_db

TEST_DB = "sqlite:///./test_schedule_nonop.db"
engine = create_engine(TEST_DB, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base.metadata.create_all(bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)

DATE = "2026-04-13"
NEXT_DATE = "2026-04-14"


@pytest.fixture(autouse=True)
def setup_teardown():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def hash_cpf(cpf: str) -> str:
    cpf_clean = re.sub(r"\D", "", cpf)
    return hashlib.sha256(cpf_clean.encode()).hexdigest()[:16]


def create_token(cpf: str, role: UserRole) -> str:
    db = TestingSessionLocal()
    user = User(
        cpf_hash=hash_cpf(cpf),
        email=f"{role.value}@test.com",
        name=f"Test {role.value}",
        password_hash=hash_password("password123"),
        role=role,
        unit=None if role == UserRole.ADMIN else "Caieiras",
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.close()
    response = client.post("/auth/login", json={"cpf": cpf, "password": "password123"})
    return response.json()["access_token"]


@pytest.fixture
def admin_token():
    return create_token("111.222.333-44", UserRole.ADMIN)


@pytest.fixture
def plantonista_token():
    return create_token("555.666.777-88", UserRole.PLANTONISTA)


def build_pair_file() -> bytes:
    """Mesma linha 7368 com Entrada (col 7) e Saida (col 9) no mesmo prefixo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "caieiras"
    ws.cell(row=4, column=1, value="1580")
    ws.cell(row=4, column=3, value="E N DA SILVA")
    # Bloco Entrada
    ws.cell(row=4, column=7, value="03:50 - 04:45")
    ws.cell(row=5, column=7, value="E/ M LIVRE - SP-02")
    ws.cell(row=6, column=7, value="L - 7368")
    ws.cell(row=7, column=7, value="JD. PINHEIROS")
    # Bloco Saida (mesma linha)
    ws.cell(row=4, column=9, value="17:00 - 18:00")
    ws.cell(row=5, column=9, value="S/ M LIVRE - SP-02")
    ws.cell(row=6, column=9, value="L - 7368")
    ws.cell(row=7, column=9, value="JD. PINHEIROS")
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def import_pair(admin_token: str) -> list[dict]:
    resp = client.post(
        f"/schedule/import?schedule_date={DATE}&replace=true",
        headers=auth(admin_token),
        files={
            "file": (
                "escala.xlsx",
                build_pair_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["imported"] == 2
    lines = client.get(
        f"/schedule/lines?schedule_date={DATE}", headers=auth(admin_token)
    ).json()
    assert len(lines) == 2
    return lines


def test_deactivate_reactivate_hides_and_restores(admin_token):
    lines = import_pair(admin_token)
    entrada = next(l for l in lines if l["direction"] == "ENTRADA")

    # Desativa (periodo) -> some da lista operacional padrao
    resp = client.post(
        f"/schedule/lines/{entrada['id']}/deactivate", headers=auth(admin_token)
    )
    assert resp.status_code == 200
    assert resp.json()["is_active"] is False

    visible = client.get(
        f"/schedule/lines?schedule_date={DATE}", headers=auth(admin_token)
    ).json()
    assert all(l["id"] != entrada["id"] for l in visible)
    assert len(visible) == 1

    # Mas a aba ADM (include_inactive) ainda ve, para reativar
    with_inactive = client.get(
        f"/schedule/lines?schedule_date={DATE}&include_inactive=true",
        headers=auth(admin_token),
    ).json()
    assert any(l["id"] == entrada["id"] and l["is_active"] is False for l in with_inactive)

    # Reativa -> volta
    resp = client.post(
        f"/schedule/lines/{entrada['id']}/reactivate", headers=auth(admin_token)
    )
    assert resp.status_code == 200
    assert resp.json()["is_active"] is True
    restored = client.get(
        f"/schedule/lines?schedule_date={DATE}", headers=auth(admin_token)
    ).json()
    assert len(restored) == 2


def test_deactivate_requires_admin(admin_token, plantonista_token):
    lines = import_pair(admin_token)
    entrada = next(l for l in lines if l["direction"] == "ENTRADA")
    resp = client.post(
        f"/schedule/lines/{entrada['id']}/deactivate", headers=auth(plantonista_token)
    )
    assert resp.status_code == 403


def test_pair_endpoint_returns_opposite_direction(admin_token):
    lines = import_pair(admin_token)
    entrada = next(l for l in lines if l["direction"] == "ENTRADA")
    saida = next(l for l in lines if l["direction"] == "SAIDA")

    pair = client.get(
        f"/schedule/lines/{entrada['id']}/pair?operation_date={DATE}",
        headers=auth(admin_token),
    ).json()
    assert len(pair) == 1
    assert pair[0]["id"] == saida["id"]
    assert pair[0]["direction"] == "SAIDA"


def test_non_operation_is_per_day_and_marks_pair(admin_token, plantonista_token):
    lines = import_pair(admin_token)
    entrada = next(l for l in lines if l["direction"] == "ENTRADA")
    saida = next(l for l in lines if l["direction"] == "SAIDA")

    # Plantonista marca "nao operar" hoje para a Entrada + a Saida (par)
    resp = client.post(
        f"/schedule/lines/{entrada['id']}/non-operation",
        headers=auth(plantonista_token),
        json={"operation_date": DATE, "also_line_ids": [saida["id"]]},
    )
    assert resp.status_code == 200, resp.text
    assert set(resp.json()["marked"]) == {entrada["id"], saida["id"]}

    # Painel (hide_non_operating) nao mostra nenhuma das duas hoje
    pending = client.get(
        f"/schedule/lines?schedule_date={DATE}&hide_non_operating=true",
        headers=auth(plantonista_token),
    ).json()
    assert pending == []

    # Busca (sem hide) mostra com flag non_operating=True
    searched = client.get(
        f"/schedule/lines?schedule_date={DATE}", headers=auth(plantonista_token)
    ).json()
    assert len(searched) == 2
    assert all(l["non_operating"] is True for l in searched)

    # No dia seguinte volta a aparecer (por dia, sem reativar nada)
    next_day = client.get(
        f"/schedule/lines?schedule_date={NEXT_DATE}&hide_non_operating=true",
        headers=auth(plantonista_token),
    ).json()
    assert len(next_day) == 2
    assert all(l["non_operating"] is False for l in next_day)

    # Desfaz a Entrada -> volta a operar hoje; Saida segue fora
    resp = client.delete(
        f"/schedule/lines/{entrada['id']}/non-operation?operation_date={DATE}",
        headers=auth(plantonista_token),
    )
    assert resp.status_code == 204
    pending2 = client.get(
        f"/schedule/lines?schedule_date={DATE}&hide_non_operating=true",
        headers=auth(plantonista_token),
    ).json()
    assert len(pending2) == 1
    assert pending2[0]["id"] == entrada["id"]


def test_plantonista_can_edit_line_inline(admin_token, plantonista_token):
    lines = import_pair(admin_token)
    entrada = next(l for l in lines if l["direction"] == "ENTRADA")

    resp = client.patch(
        f"/schedule/lines/{entrada['id']}",
        headers=auth(plantonista_token),
        json={"driver_name": "NOVO MOTORISTA", "status": entrada["status"]},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["driver_name"] == "NOVO MOTORISTA"
    # Status preservado (nao virou "alterada" nem sumiu dos pendentes)
    assert resp.json()["status"] == entrada["status"]


def _preview(admin_token: str, filename: str, schedule_date: str) -> dict:
    return client.post(
        f"/schedule/import/preview?schedule_date={schedule_date}",
        headers=auth(admin_token),
        files={
            "file": (
                filename,
                build_pair_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    ).json()


def test_preview_avisa_coexistencia_de_vigencia(admin_token):
    import_pair(admin_token)  # importa "escala.xlsx" na vigencia DATE

    # Mesmo nome + mesma vigencia -> vai substituir
    same = _preview(admin_token, "escala.xlsx", DATE)
    assert same["will_replace"] is True
    assert same["existing_other_files"] == []

    # Nome diferente + mesma vigencia -> coexistencia (duplicaria)
    other = _preview(admin_token, "OUTRA ESCALA.xlsx", DATE)
    assert other["will_replace"] is False
    assert "escala.xlsx" in other["existing_other_files"]

    # Nome diferente + OUTRA vigencia -> sem coexistencia
    other_date = _preview(admin_token, "OUTRA ESCALA.xlsx", NEXT_DATE)
    assert other_date["existing_other_files"] == []
